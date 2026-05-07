from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
CSV_EXTENSIONS = {".csv"}
LEGACY_EXCEL_EXTENSIONS = {".xls"}


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile local inbox spreadsheet files.")
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument("--inbox-root", type=Path)
    parser.add_argument("--state-root", type=Path)
    args = parser.parse_args()

    repo_root = args.repo_root.resolve()
    inbox_root = args.inbox_root or repo_root / "data" / "inbox"
    state_root = args.state_root or repo_root / "data" / "state"
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    state_root.mkdir(parents=True, exist_ok=True)

    profiles = [profile_file(path, repo_root) for path in sorted(inbox_root.rglob("*")) if path.is_file()]
    summary = summarize_profiles(profiles)
    snapshot = {
        "run_id": run_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "inbox_root": str(inbox_root.relative_to(repo_root)),
        "summary": summary,
        "files": profiles,
    }

    run_path = state_root / f"inbox_profile_{run_id}.json"
    latest_path = state_root / "inbox_profile_latest.json"
    run_path.write_text(json.dumps(snapshot, indent=2, sort_keys=True), encoding="utf-8")
    latest_path.write_text(json.dumps(snapshot, indent=2, sort_keys=True), encoding="utf-8")

    print(f"profile={run_path.relative_to(repo_root)}")
    print(f"latest={latest_path.relative_to(repo_root)}")
    print(
        "summary "
        f"files={summary['file_count']} "
        f"excel_opened={summary['excel_opened']} "
        f"csv_opened={summary['csv_opened']} "
        f"unsupported={summary['unsupported_count']} "
        f"errors={summary['error_count']}"
    )
    print(f"extensions={summary['extensions']}")
    print(f"folders={summary['folders']}")
    return 1 if summary["error_count"] else 0


def profile_file(path: Path, repo_root: Path) -> dict[str, Any]:
    extension = path.suffix.lower()
    base = {
        "local_path": str(path.relative_to(repo_root)),
        "name": path.name,
        "extension": extension,
        "size": path.stat().st_size,
    }

    try:
        if extension in EXCEL_EXTENSIONS:
            return {**base, **profile_workbook(path)}
        if extension in CSV_EXTENSIONS:
            return {**base, **profile_csv(path)}
        if extension in LEGACY_EXCEL_EXTENSIONS:
            return {**base, "kind": "legacy_excel", "status": "unsupported_without_xlrd"}
        return {**base, "kind": "unknown", "status": "unsupported_extension"}
    except Exception as exc:  # noqa: BLE001 - record file-level profiling failures.
        return {**base, "kind": "unknown", "status": "error", "error": str(exc)}


def profile_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header = find_header_like_row(sheet)
            sheets.append(
                {
                    "name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "header_row_index": header["row_index"],
                    "header_non_empty_cells": header["non_empty_cells"],
                    "header_values": header["values"],
                }
            )
        return {
            "kind": "excel",
            "status": "profiled",
            "sheet_count": len(sheets),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets,
        }
    finally:
        workbook.close()


def profile_csv(path: Path) -> dict[str, Any]:
    sample_rows = []
    encoding = "utf-8-sig"
    try:
        with path.open(newline="", encoding=encoding) as file:
            reader = csv.reader(file)
            for index, row in enumerate(reader, 1):
                sample_rows.append(row)
                if index >= 10:
                    break
    except UnicodeDecodeError:
        encoding = "latin-1"
        with path.open(newline="", encoding=encoding) as file:
            reader = csv.reader(file)
            for index, row in enumerate(reader, 1):
                sample_rows.append(row)
                if index >= 10:
                    break

    header = first_header_like_row(sample_rows)
    return {
        "kind": "csv",
        "status": "profiled",
        "encoding": encoding,
        "header_row_index": header["row_index"],
        "header_non_empty_cells": header["non_empty_cells"],
        "header_values": header["values"],
    }


def find_header_like_row(sheet) -> dict[str, Any]:
    rows = []
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        rows.append(list(row or []))
    return first_header_like_row(rows)


def first_header_like_row(rows: list[list[Any]]) -> dict[str, Any]:
    best = {"row_index": None, "non_empty_cells": 0, "values": []}
    for index, row in enumerate(rows, 1):
        values = [normalize_cell(value) for value in row]
        values = [value for value in values if value]
        if len(values) > best["non_empty_cells"]:
            best = {
                "row_index": index,
                "non_empty_cells": len(values),
                "values": values[:40],
            }
        if len(values) >= 3:
            return best
    return best


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > 120:
        return text[:117] + "..."
    return text


def summarize_profiles(profiles: list[dict[str, Any]]) -> dict[str, Any]:
    extensions = Counter(row["extension"] for row in profiles)
    folders: defaultdict[str, int] = defaultdict(int)
    for row in profiles:
        parts = Path(row["local_path"]).parts
        folder = "/".join(parts[:4]) if len(parts) >= 4 else str(Path(row["local_path"]).parent)
        folders[folder] += 1

    statuses = Counter(row.get("status", "") for row in profiles)
    kinds = Counter(row.get("kind", "") for row in profiles)
    return {
        "file_count": len(profiles),
        "extensions": dict(sorted(extensions.items())),
        "folders": dict(sorted(folders.items(), key=lambda item: (-item[1], item[0]))),
        "statuses": dict(sorted(statuses.items())),
        "kinds": dict(sorted(kinds.items())),
        "excel_opened": kinds.get("excel", 0),
        "csv_opened": kinds.get("csv", 0),
        "unsupported_count": statuses.get("unsupported_without_xlrd", 0) + statuses.get("unsupported_extension", 0),
        "error_count": statuses.get("error", 0),
    }


if __name__ == "__main__":
    raise SystemExit(main())
