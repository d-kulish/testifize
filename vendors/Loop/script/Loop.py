from __future__ import annotations

import csv
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[2]

INPUT_SCHEMA_PATH = SCRIPT_DIR / "input_schema.json"
FINAL_LOOP_PATH = PROJECT_ROOT / "final" / "Loop.csv"
OUTPUT_PATH = SCRIPT_DIR / "Loop.csv"
REPORT_PATH = SCRIPT_DIR / "Loop_comparison_report.md"

SPEND_TOLERANCE = Decimal("0.000001")
IMPRESSION_TOLERANCE = Decimal("0")


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid Date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def build_header_map(input_schema: dict) -> dict[str, int]:
    return {
        column["name"]: column_letter_to_index(column["letter"])
        for column in input_schema["header"]["columns"]
    }


def validate_header(sheet, input_schema: dict, header_map: dict[str, int]) -> None:
    header_row = input_schema["header"]["row"]
    expected = {column["name"] for column in input_schema["header"]["columns"]}
    actual = {
        name
        for name, column_index in header_map.items()
        if sheet.cell(header_row, column_index).value == name
    }
    missing = sorted(expected - actual)
    if missing:
        raise ValueError(f"Input schema mismatch on header row {header_row}: {missing}")


def source_rows_and_subtotal() -> tuple[list[dict[str, object]], dict[str, Decimal] | None]:
    input_schema = load_json(INPUT_SCHEMA_PATH)
    source_path = (SCRIPT_DIR / input_schema["source_file"]).resolve()
    selected = input_schema["selected_columns"]
    defaults = input_schema["output_defaults"]

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook[input_schema["sheet_name"]]
    header_map = build_header_map(input_schema)
    validate_header(sheet, input_schema, header_map)

    rows: list[dict[str, object]] = []
    subtotal: dict[str, Decimal] | None = None

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=input_schema["header"]["row"] + 1,
            max_row=sheet.max_row,
            values_only=True,
        ),
        start=input_schema["header"]["row"] + 1,
    ):
        date_value = row[header_map[selected["date"]] - 1]
        impression_value = row[header_map[selected["impressions"]] - 1]
        spend_value = row[header_map[selected["spend"]] - 1]

        if date_value in (None, ""):
            if impression_value not in (None, "") and spend_value not in (None, ""):
                subtotal = {
                    "Impressions": parse_decimal(
                        impression_value,
                        row_number,
                        selected["impressions"],
                    ),
                    "Spend": parse_decimal(spend_value, row_number, selected["spend"]),
                }
            continue

        rows.append(
            {
                "Date": parse_date(date_value, row_number),
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Channel": defaults["Channel"],
                "Platform": defaults["Platform"],
                "Spend": parse_decimal(spend_value, row_number, selected["spend"]),
                "Impressions": parse_decimal(
                    impression_value,
                    row_number,
                    selected["impressions"],
                ),
                "Data_Grain": defaults["Data_Grain"],
                "Processed_At": defaults["Processed_At"],
                "Source_File": defaults["Source_File"],
            }
        )

    if not rows:
        raise ValueError(f"No rows were parsed from {source_path}.")
    return rows, subtotal


def load_final_columns() -> list[str]:
    with FINAL_LOOP_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        return next(reader)


def write_csv(rows: list[dict[str, object]]) -> None:
    fieldnames = load_final_columns()
    with OUTPUT_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            output_row = dict(row)
            output_row["Spend"] = decimal_text(row["Spend"])
            output_row["Impressions"] = decimal_text(row["Impressions"])
            writer.writerow(output_row)


def load_generated_rows() -> list[dict[str, str | Decimal]]:
    with OUTPUT_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [normalize_row(row, row_number) for row_number, row in enumerate(reader, start=2)]


def load_final_feb_rows() -> list[dict[str, str | Decimal]]:
    input_schema = load_json(INPUT_SCHEMA_PATH)
    start_date = input_schema["comparison_window"]["start_date"]
    end_date = input_schema["comparison_window"]["end_date"]
    rows = []
    with FINAL_LOOP_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            if start_date <= row["Date"] <= end_date:
                rows.append(normalize_row(row, row_number))
    return rows


def normalize_row(row: dict[str, str], row_number: int) -> dict[str, str | Decimal]:
    return {
        "Date": row["Date"],
        "Vendor": row["Vendor"],
        "Brand": row["Brand"],
        "Channel": row["Channel"],
        "Platform": row["Platform"],
        "Spend": parse_decimal(row["Spend"], row_number, "Spend"),
        "Impressions": parse_decimal(row["Impressions"], row_number, "Impressions"),
        "Data_Grain": row["Data_Grain"],
        "Processed_At": row["Processed_At"],
        "Source_File": row["Source_File"],
    }


def summarize(rows: list[dict[str, str | Decimal]]) -> dict[str, object]:
    dates = [row["Date"] for row in rows]
    return {
        "rows": len(rows),
        "min_date": min(dates) if dates else "",
        "max_date": max(dates) if dates else "",
        "spend": sum((row["Spend"] for row in rows), Decimal("0")),
        "impressions": sum((row["Impressions"] for row in rows), Decimal("0")),
        "vendors": sorted({row["Vendor"] for row in rows}),
        "brands": sorted({row["Brand"] for row in rows}),
        "channels": sorted({row["Channel"] for row in rows}),
        "platforms": sorted({row["Platform"] for row in rows}),
        "data_grains": sorted({row["Data_Grain"] for row in rows}),
        "processed_at": sorted({row["Processed_At"] for row in rows}),
        "source_files": sorted({row["Source_File"] for row in rows}),
    }


def compare(
    generated_rows: list[dict[str, str | Decimal]],
    final_rows: list[dict[str, str | Decimal]],
    subtotal: dict[str, Decimal] | None,
) -> dict[str, object]:
    generated_by_date = {row["Date"]: row for row in generated_rows}
    final_by_date = {row["Date"]: row for row in final_rows}
    generated_dates = set(generated_by_date)
    final_dates = set(final_by_date)

    differing_rows = []
    max_spend_diff = Decimal("0")
    max_impression_diff = Decimal("0")
    for current_date in sorted(generated_dates & final_dates):
        generated = generated_by_date[current_date]
        final = final_by_date[current_date]
        spend_diff = generated["Spend"] - final["Spend"]
        impression_diff = generated["Impressions"] - final["Impressions"]
        max_spend_diff = max(max_spend_diff, abs(spend_diff))
        max_impression_diff = max(max_impression_diff, abs(impression_diff))
        if abs(spend_diff) > SPEND_TOLERANCE or abs(impression_diff) > IMPRESSION_TOLERANCE:
            differing_rows.append(
                {
                    "Date": current_date,
                    "Generated Spend": generated["Spend"],
                    "Final Spend": final["Spend"],
                    "Spend Diff": spend_diff,
                    "Generated Impressions": generated["Impressions"],
                    "Final Impressions": final["Impressions"],
                    "Impressions Diff": impression_diff,
                }
            )

    generated_summary = summarize(generated_rows)
    final_summary = summarize(final_rows)
    subtotal_diff = None
    if subtotal is not None:
        subtotal_diff = {
            "Spend": subtotal["Spend"] - generated_summary["spend"],
            "Impressions": subtotal["Impressions"] - generated_summary["impressions"],
            "Subtotal Spend": subtotal["Spend"],
            "Subtotal Impressions": subtotal["Impressions"],
        }

    value_checks = {
        "Vendor": generated_summary["vendors"] == final_summary["vendors"],
        "Brand": generated_summary["brands"] == final_summary["brands"],
        "Channel": generated_summary["channels"] == final_summary["channels"],
        "Platform": generated_summary["platforms"] == final_summary["platforms"],
        "Data_Grain": generated_summary["data_grains"] == final_summary["data_grains"],
    }

    return {
        "generated_summary": generated_summary,
        "final_summary": final_summary,
        "missing_dates": sorted(final_dates - generated_dates),
        "extra_dates": sorted(generated_dates - final_dates),
        "differing_rows": differing_rows,
        "max_spend_diff": max_spend_diff,
        "max_impression_diff": max_impression_diff,
        "subtotal_diff": subtotal_diff,
        "value_checks": value_checks,
    }


def write_report(comparison: dict[str, object]) -> None:
    generated = comparison["generated_summary"]
    final = comparison["final_summary"]
    differing_rows = comparison["differing_rows"]
    subtotal_diff = comparison["subtotal_diff"]

    lines = [
        "# Loop February 2026 Comparison Report",
        "",
        "## Summary",
        "",
        "| Metric | Generated Loop.csv | final/Loop.csv Feb 2026 |",
        "|---|---:|---:|",
        f"| Rows | {generated['rows']} | {final['rows']} |",
        f"| Date range | {generated['min_date']} to {generated['max_date']} | "
        f"{final['min_date']} to {final['max_date']} |",
        f"| Total spend | {decimal_text(generated['spend'])} | {decimal_text(final['spend'])} |",
        f"| Total impressions | {decimal_text(generated['impressions'])} | "
        f"{decimal_text(final['impressions'])} |",
        "",
        "## Date Coverage",
        "",
        f"- Missing dates from generated output: {len(comparison['missing_dates'])}",
        f"- Extra dates in generated output: {len(comparison['extra_dates'])}",
        "",
        "## Comparable Value Checks",
        "",
        "| Check | Result |",
        "|---|---|",
    ]

    for label, matched in comparison["value_checks"].items():
        lines.append(f"| {label} | {'PASS' if matched else 'DIFF'} |")

    lines.extend(
        [
            "",
            "## Difference Checks",
            "",
            f"- Max absolute daily spend difference: {comparison['max_spend_diff']}",
            f"- Max absolute daily impression difference: {comparison['max_impression_diff']}",
            f"- Material differing dates: {len(differing_rows)}",
            "",
            "## Metadata",
            "",
            f"- Generated `Source_File`: {', '.join(generated['source_files'])}",
            f"- Final Feb `Source_File`: {', '.join(final['source_files'])}",
            f"- Generated `Processed_At`: {', '.join(generated['processed_at'])}",
            f"- Final Feb `Processed_At`: {', '.join(final['processed_at'])}",
            "",
        ]
    )

    if subtotal_diff is not None:
        lines.extend(
            [
                "## Source Subtotal Note",
                "",
                f"- Dated source rows spend total: {decimal_text(generated['spend'])}",
                f"- Source subtotal row spend: {decimal_text(subtotal_diff['Subtotal Spend'])}",
                f"- Source subtotal row spend difference: {decimal_text(subtotal_diff['Spend'])}",
                f"- Dated source rows impressions total: {decimal_text(generated['impressions'])}",
                f"- Source subtotal row impressions: {decimal_text(subtotal_diff['Subtotal Impressions'])}",
                f"- Source subtotal row impressions difference: {decimal_text(subtotal_diff['Impressions'])}",
                "",
            ]
        )

    if differing_rows:
        lines.extend(
            [
                "## Sample Differing Dates",
                "",
                "| Date | Generated Spend | Final Spend | Spend Diff | Generated Impressions | Final Impressions | Impressions Diff |",
                "|---|---:|---:|---:|---:|---:|---:|",
            ]
        )
        for row in differing_rows[:10]:
            lines.append(
                f"| {row['Date']} | {decimal_text(row['Generated Spend'])} | "
                f"{decimal_text(row['Final Spend'])} | {decimal_text(row['Spend Diff'])} | "
                f"{decimal_text(row['Generated Impressions'])} | "
                f"{decimal_text(row['Final Impressions'])} | "
                f"{decimal_text(row['Impressions Diff'])} |"
            )
        lines.append("")
    else:
        lines.extend(["No material differences were found for February 2026 dated rows.", ""])

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    rows, subtotal = source_rows_and_subtotal()
    write_csv(rows)

    generated_rows = load_generated_rows()
    final_rows = load_final_feb_rows()
    comparison = compare(generated_rows, final_rows, subtotal)
    write_report(comparison)

    print(f"Wrote {len(generated_rows)} rows to {OUTPUT_PATH}")
    print(f"Wrote comparison report to {REPORT_PATH}")


if __name__ == "__main__":
    main()
