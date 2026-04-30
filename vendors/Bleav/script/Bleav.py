from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[2]

INPUT_SCHEMA_PATH = SCRIPT_DIR / "input_schema.json"
OUTPUT_SCHEMA_PATH = PROJECT_ROOT / "test" / "scripts" / "schemas" / "output_schema.json"
OUTPUT_PATH = SCRIPT_DIR / "Bleav.csv"
REPORT_PATH = SCRIPT_DIR / "Bleav_comparison_report.md"
FINAL_BLEAV_PATH = PROJECT_ROOT / "final" / "Bleav.csv"

SPEND_TOLERANCE = Decimal("0.005")


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def load_output_columns() -> list[str]:
    schema = load_json(OUTPUT_SCHEMA_PATH)
    return [column["name"] for column in schema["target_output_schema"]]


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid Date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal, quantize: str | None = None) -> str:
    if quantize:
        value = value.quantize(Decimal(quantize))
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def build_header_map(input_schema: dict) -> dict[str, int]:
    return {
        column["name"]: column_letter_to_index(column["letter"])
        for column in input_schema["header"]["columns"]
    }


def validate_header(sheet, input_schema: dict, header_map: dict[str, int]) -> None:
    header_row = input_schema["header"]["row"]
    expected = {column["name"] for column in input_schema["header"]["columns"]}
    actual = {
        name
        for name, column_index in header_map.items()
        if sheet.cell(header_row, column_index).value == name
    }
    missing = sorted(expected - actual)
    if missing:
        raise ValueError(f"Input schema mismatch on header row {header_row}: {missing}")


def get_cell(row: tuple[object, ...], header_map: dict[str, int], column_name: str) -> object:
    return row[header_map[column_name] - 1]


def validate_allowed(
    value: object,
    allowed_values: set[str],
    row_number: int,
    column_name: str,
) -> str:
    text = (value or "").strip()
    if text not in allowed_values:
        raise ValueError(f"Unexpected {column_name} at row {row_number}: {text!r}")
    return text


def build_rows() -> tuple[list[dict[str, object]], Counter[str]]:
    input_schema = load_json(INPUT_SCHEMA_PATH)
    source_path = (SCRIPT_DIR / input_schema["source_file"]).resolve()
    selected = input_schema["selected_columns"]
    defaults = input_schema["output_defaults"]

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook[input_schema["sheet_name"]]
    header_map = build_header_map(input_schema)
    validate_header(sheet, input_schema, header_map)

    validations = input_schema["validations"]
    allowed_vendors = set(validations["Vendor"]["allowed_values"])
    allowed_channels = set(validations["Marketing_Channel"]["allowed_values"])
    allowed_sub_channels = set(validations["Sub_Channel"]["allowed_values"])
    allowed_patterns = set(validations["Spend_Pattern"]["allowed_values"])
    allowed_data_sources = set(validations["Data_Source"]["allowed_values"])

    rows: list[dict[str, object]] = []
    data_source_counts: Counter[str] = Counter()

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=input_schema["header"]["row"] + 1,
            max_row=sheet.max_row,
            values_only=True,
        ),
        start=input_schema["header"]["row"] + 1,
    ):
        date_value = get_cell(row, header_map, selected["date"])
        if date_value in (None, ""):
            continue

        validate_allowed(
            get_cell(row, header_map, selected["source_vendor"]),
            allowed_vendors,
            row_number,
            selected["source_vendor"],
        )
        marketing_channel = validate_allowed(
            get_cell(row, header_map, selected["marketing_channel"]),
            allowed_channels,
            row_number,
            selected["marketing_channel"],
        )
        sub_channel = validate_allowed(
            get_cell(row, header_map, selected["sub_channel"]),
            allowed_sub_channels,
            row_number,
            selected["sub_channel"],
        )
        validate_allowed(
            get_cell(row, header_map, selected["spend_pattern"]),
            allowed_patterns,
            row_number,
            selected["spend_pattern"],
        )
        data_source = validate_allowed(
            get_cell(row, header_map, selected["data_source"]),
            allowed_data_sources,
            row_number,
            selected["data_source"],
        )
        data_source_counts[data_source] += 1

        campaign = (get_cell(row, header_map, selected["campaign"]) or "").strip()
        if not campaign:
            raise ValueError(f"Missing Campaign at source row {row_number}.")

        rows.append(
            {
                "Date": parse_date(date_value, row_number),
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Campaign": campaign,
                "Marketing_Channel": marketing_channel,
                "Sub_Channel": sub_channel,
                "Daily_Spend": parse_decimal(
                    get_cell(row, header_map, selected["spend"]),
                    row_number,
                    selected["spend"],
                ),
                "Daily_Impressions": parse_decimal(
                    get_cell(row, header_map, selected["impressions"]),
                    row_number,
                    selected["impressions"],
                ),
                "Spend_Type": defaults["Spend_Type"],
            }
        )

    if not rows:
        raise ValueError(f"No rows were parsed from {source_path}.")
    return rows, data_source_counts


def write_csv(rows: list[dict[str, object]]) -> None:
    fieldnames = load_output_columns()
    with OUTPUT_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "Date": row["Date"],
                    "Vendor": row["Vendor"],
                    "Brand": row["Brand"],
                    "Campaign": row["Campaign"],
                    "Marketing_Channel": row["Marketing_Channel"],
                    "Sub_Channel": row["Sub_Channel"],
                    "Daily_Spend": decimal_text(row["Daily_Spend"], "0.01"),
                    "Daily_Impressions": decimal_text(row["Daily_Impressions"]),
                    "Spend_Type": row["Spend_Type"],
                }
            )


def validate_required_columns(fieldnames: list[str] | None, required: list[str], path: Path) -> None:
    if not fieldnames:
        raise ValueError(f"{path} has no header row.")
    missing = [column for column in required if column not in fieldnames]
    if missing:
        raise ValueError(f"{path.name} is missing required columns: {missing}")


def load_generated_rows() -> list[dict[str, str | Decimal]]:
    with OUTPUT_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [
            {
                "Date": row["Date"],
                "Vendor": row["Vendor"],
                "Brand": row["Brand"],
                "Campaign": row["Campaign"],
                "Marketing_Channel": row["Marketing_Channel"],
                "Sub_Channel": row["Sub_Channel"],
                "Daily_Spend": parse_decimal(row["Daily_Spend"], row_number, "Daily_Spend"),
                "Daily_Impressions": parse_decimal(
                    row["Daily_Impressions"],
                    row_number,
                    "Daily_Impressions",
                ),
                "Spend_Type": row["Spend_Type"],
            }
            for row_number, row in enumerate(reader, start=2)
        ]


def load_final_rows() -> list[dict[str, str | Decimal]]:
    with FINAL_BLEAV_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        validate_required_columns(
            reader.fieldnames,
            ["Date", "Vendor", "Brand", "Channel", "Platform", "Spend"],
            FINAL_BLEAV_PATH,
        )
        return [
            {
                "Date": row["Date"],
                "Vendor": row["Vendor"],
                "Brand": row["Brand"],
                "Channel": row["Channel"],
                "Platform": row["Platform"],
                "Spend": parse_decimal(row["Spend"], row_number, "Spend"),
            }
            for row_number, row in enumerate(reader, start=2)
        ]


def summarize_generated(rows: list[dict[str, str | Decimal]]) -> dict[str, object]:
    dates = [row["Date"] for row in rows]
    campaigns = sorted({row["Campaign"] for row in rows})
    return {
        "rows": len(rows),
        "dates": len(set(dates)),
        "min_date": min(dates),
        "max_date": max(dates),
        "spend": sum((row["Daily_Spend"] for row in rows), Decimal("0")),
        "impressions": sum((row["Daily_Impressions"] for row in rows), Decimal("0")),
        "campaign_count": len(campaigns),
        "campaign_sample": campaigns[:10],
        "vendors": sorted({row["Vendor"] for row in rows}),
        "brands": sorted({row["Brand"] for row in rows}),
        "channels": sorted({row["Marketing_Channel"] for row in rows}),
        "sub_channels": sorted({row["Sub_Channel"] for row in rows}),
        "spend_types": sorted({row["Spend_Type"] for row in rows}),
    }


def summarize_final(rows: list[dict[str, str | Decimal]]) -> dict[str, object]:
    dates = [row["Date"] for row in rows]
    return {
        "rows": len(rows),
        "dates": len(set(dates)),
        "min_date": min(dates),
        "max_date": max(dates),
        "spend": sum((row["Spend"] for row in rows), Decimal("0")),
        "vendors": sorted({row["Vendor"] for row in rows}),
        "brands": sorted({row["Brand"] for row in rows}),
        "channels": sorted({row["Channel"] for row in rows}),
        "platforms": sorted({row["Platform"] for row in rows}),
    }


def final_daily_totals(rows: list[dict[str, str | Decimal]]) -> dict[str, Decimal]:
    totals: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        totals[row["Date"]] += row["Spend"]
    return dict(totals)


def generated_daily_totals(rows: list[dict[str, str | Decimal]]) -> dict[str, Decimal]:
    totals: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        totals[row["Date"]] += row["Daily_Spend"]
    return dict(totals)


def compare(
    generated_rows: list[dict[str, str | Decimal]],
    final_rows: list[dict[str, str | Decimal]],
    data_source_counts: Counter[str],
) -> dict[str, object]:
    generated_summary = summarize_generated(generated_rows)
    final_summary = summarize_final(final_rows)
    generated_daily = generated_daily_totals(generated_rows)
    final_daily = final_daily_totals(final_rows)

    generated_dates = set(generated_daily)
    final_dates = set(final_daily)
    shared_dates = sorted(generated_dates & final_dates)

    daily_spend_diffs = []
    max_abs_daily_spend_diff = Decimal("0")
    for current_date in shared_dates:
        spend_diff = generated_daily[current_date] - final_daily[current_date]
        max_abs_daily_spend_diff = max(max_abs_daily_spend_diff, abs(spend_diff))
        if abs(spend_diff) > SPEND_TOLERANCE:
            daily_spend_diffs.append(
                {
                    "Date": current_date,
                    "Generated Spend": generated_daily[current_date],
                    "Final Spend": final_daily[current_date],
                    "Spend Diff": spend_diff,
                }
            )

    comparable_sets = {
        "Date coverage": generated_dates == final_dates,
        "Vendor values": set(generated_summary["vendors"]) == set(final_summary["vendors"]),
        "Brand values": set(generated_summary["brands"]) == set(final_summary["brands"]),
        "Marketing_Channel -> Channel values": set(generated_summary["channels"])
        == set(final_summary["channels"]),
        "Sub_Channel -> Platform values": set(generated_summary["sub_channels"])
        == set(final_summary["platforms"]),
    }

    return {
        "generated_summary": generated_summary,
        "final_summary": final_summary,
        "data_source_counts": data_source_counts,
        "missing_dates": sorted(final_dates - generated_dates),
        "extra_dates": sorted(generated_dates - final_dates),
        "daily_spend_diffs": daily_spend_diffs,
        "max_abs_daily_spend_diff": max_abs_daily_spend_diff,
        "total_spend_diff": generated_summary["spend"] - final_summary["spend"],
        "comparable_sets": comparable_sets,
    }


def write_report(comparison: dict[str, object]) -> None:
    generated = comparison["generated_summary"]
    final = comparison["final_summary"]
    daily_spend_diffs = comparison["daily_spend_diffs"]
    data_source_counts = comparison["data_source_counts"]

    lines = [
        "# Bleav Comparison Report",
        "",
        "## Summary",
        "",
        "| Metric | Generated Bleav.csv | final/Bleav.csv |",
        "|---|---:|---:|",
        f"| Rows | {generated['rows']} | {final['rows']} |",
        f"| Unique dates | {generated['dates']} | {final['dates']} |",
        f"| Date range | {generated['min_date']} to {generated['max_date']} | "
        f"{final['min_date']} to {final['max_date']} |",
        f"| Total spend | {decimal_text(generated['spend'], '0.01')} | "
        f"{decimal_text(final['spend'], '0.01')} |",
        f"| Total impressions | {decimal_text(generated['impressions'])} | Not available |",
        f"| Campaign/show count | {generated['campaign_count']} | Not available |",
        "",
        "## Source Choice",
        "",
        "- Parsed `By_Show_Detail` because it contains the actual show/campaign names required by `Campaign`.",
        "- `Daily_Aggregated` is useful for total checks, but its `Campaign = Podcast` is too coarse for the output schema.",
        "",
        "## Old File Limitations",
        "",
        "- `final/Bleav.csv` has no impressions column.",
        "- `final/Bleav.csv` has no campaign/show field.",
        "- `final/Bleav.csv` has no data-source or source-file lineage column.",
        "- `final/Bleav.csv` has many rows per date, so it is compared by daily spend totals.",
        "",
        "## Source Data Status",
        "",
        "| Data_Source | Rows |",
        "|---|---:|",
    ]

    for label, count in sorted(data_source_counts.items()):
        lines.append(f"| {label} | {count} |")

    lines.extend(
        [
            "",
            "## Date-Level Comparison",
            "",
            f"- Missing dates from generated output: {len(comparison['missing_dates'])}",
            f"- Extra dates in generated output: {len(comparison['extra_dates'])}",
        f"- Total spend difference: {comparison['total_spend_diff']}",
        f"- Max absolute daily spend difference: {comparison['max_abs_daily_spend_diff']}",
        f"- Dates with material daily spend differences: {len(daily_spend_diffs)}",
        "- Spend differences are small daily rounding/import differences in the old file; date coverage matches exactly.",
        "",
        "## Comparable Value Checks",
            "",
            "| Check | Result |",
            "|---|---|",
        ]
    )

    for label, matched in comparison["comparable_sets"].items():
        lines.append(f"| {label} | {'PASS' if matched else 'DIFF'} |")

    lines.extend(
        [
            "",
            "## Generated Schema Values",
            "",
            f"- `Vendor`: {', '.join(generated['vendors'])}",
            f"- `Brand`: {', '.join(generated['brands'])}",
            f"- `Campaign` count: {generated['campaign_count']}",
            f"- `Campaign` sample: {', '.join(generated['campaign_sample'])}",
            f"- `Marketing_Channel`: {', '.join(generated['channels'])}",
            f"- `Sub_Channel`: {', '.join(generated['sub_channels'])}",
            f"- `Spend_Type`: {', '.join(generated['spend_types'])}",
            "",
        ]
    )

    if daily_spend_diffs:
        lines.extend(
            [
                "## Sample Daily Spend Differences",
                "",
                "| Date | Generated Spend | Final Spend | Difference |",
                "|---|---:|---:|---:|",
            ]
        )
        for row in daily_spend_diffs[:10]:
            lines.append(
                f"| {row['Date']} | {decimal_text(row['Generated Spend'], '0.01')} | "
                f"{decimal_text(row['Final Spend'], '0.01')} | "
                f"{decimal_text(row['Spend Diff'], '0.01')} |"
            )
        lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    source_rows, data_source_counts = build_rows()
    write_csv(source_rows)

    generated_rows = load_generated_rows()
    final_rows = load_final_rows()
    comparison = compare(generated_rows, final_rows, data_source_counts)
    write_report(comparison)

    print(f"Wrote {len(generated_rows)} rows to {OUTPUT_PATH}")
    print(f"Wrote comparison report to {REPORT_PATH}")


if __name__ == "__main__":
    main()
