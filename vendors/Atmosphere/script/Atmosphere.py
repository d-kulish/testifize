from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[2]

INPUT_SCHEMA_PATH = SCRIPT_DIR / "input_schema.json"
OUTPUT_SCHEMA_PATH = PROJECT_ROOT / "test" / "scripts" / "schemas" / "output_schema.json"
OUTPUT_PATH = SCRIPT_DIR / "Atmosphere.csv"
REPORT_PATH = SCRIPT_DIR / "Atmosphere_comparison_report.md"
FINAL_ATMOSPHERE_PATH = PROJECT_ROOT / "final" / "Atmosphere.csv"

SOURCE_TOKEN = "Atmosphere_Utah_BetOnline_Football_Dec_24_Mar_25"
SPEND_TOLERANCE = Decimal("0.005")
IMPRESSION_TOLERANCE = Decimal("0.000001")


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def load_output_columns() -> list[str]:
    schema = load_json(OUTPUT_SCHEMA_PATH)
    return [column["name"] for column in schema["target_output_schema"]]


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid Date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal, quantize: str | None = None) -> str:
    if quantize:
        value = value.quantize(Decimal(quantize))
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def build_header_map(input_schema: dict) -> dict[str, int]:
    header = input_schema["header"]
    return {
        column["name"]: column_letter_to_index(column["letter"])
        for column in header["columns"]
    }


def validate_header(sheet, input_schema: dict, header_map: dict[str, int]) -> None:
    row_number = input_schema["header"]["row"]
    expected = {column["name"] for column in input_schema["header"]["columns"]}
    actual = {
        name
        for name, column_index in header_map.items()
        if sheet.cell(row_number, column_index).value == name
    }
    missing = sorted(expected - actual)
    if missing:
        raise ValueError(f"Input schema mismatch on header row {row_number}: {missing}")


def build_daily_rows() -> list[dict[str, object]]:
    input_schema = load_json(INPUT_SCHEMA_PATH)
    source_path = (SCRIPT_DIR / input_schema["source_file"]).resolve()
    selected = input_schema["selected_columns"]
    defaults = input_schema["output_defaults"]
    allowed_brands = set(input_schema["validations"]["Advertiser"]["allowed_values"])

    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[input_schema["sheet_name"]]
    header_map = build_header_map(input_schema)
    validate_header(sheet, input_schema, header_map)

    daily_spend: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    daily_impressions: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))

    for row_number in range(input_schema["header"]["row"] + 1, sheet.max_row + 1):
        date_value = sheet.cell(row_number, header_map[selected["date"]]).value
        spend_value = sheet.cell(row_number, header_map[selected["spend"]]).value
        impression_value = sheet.cell(row_number, header_map[selected["impressions"]]).value
        brand_value = sheet.cell(row_number, header_map[selected["brand"]]).value

        if date_value in (None, ""):
            continue

        brand = (brand_value or "").strip()
        if brand not in allowed_brands:
            raise ValueError(f"Unexpected advertiser at row {row_number}: {brand!r}")

        current_date = parse_date(date_value, row_number)
        daily_spend[current_date] += parse_decimal(spend_value, row_number, selected["spend"])
        daily_impressions[current_date] += parse_decimal(
            impression_value,
            row_number,
            selected["impressions"],
        )

    rows: list[dict[str, object]] = []
    for current_date in sorted(daily_spend):
        rows.append(
            {
                "Date": current_date,
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Campaign": defaults["Campaign"],
                "Marketing_Channel": defaults["Marketing_Channel"],
                "Sub_Channel": defaults["Sub_Channel"],
                "Daily_Spend": daily_spend[current_date],
                "Daily_Impressions": daily_impressions[current_date],
                "Spend_Type": defaults["Spend_Type"],
            }
        )

    if not rows:
        raise ValueError(f"No rows were parsed from {source_path}.")
    return rows


def write_csv(rows: list[dict[str, object]]) -> None:
    fieldnames = load_output_columns()
    with OUTPUT_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            output_row = dict(row)
            output_row["Daily_Spend"] = decimal_text(row["Daily_Spend"], "0.01")
            output_row["Daily_Impressions"] = decimal_text(row["Daily_Impressions"])
            writer.writerow(output_row)


def validate_required_columns(fieldnames: list[str] | None, required: list[str], path: Path) -> None:
    if not fieldnames:
        raise ValueError(f"{path} has no header row.")
    missing = [column for column in required if column not in fieldnames]
    if missing:
        raise ValueError(f"{path.name} is missing required columns: {missing}")


def load_generated_daily() -> dict[str, dict[str, str | Decimal]]:
    daily: dict[str, dict[str, str | Decimal]] = {}
    with OUTPUT_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            daily[row["Date"]] = {
                "Date": row["Date"],
                "Vendor": row["Vendor"],
                "Brand": row["Brand"],
                "Campaign": row["Campaign"],
                "Marketing_Channel": row["Marketing_Channel"],
                "Sub_Channel": row["Sub_Channel"],
                "Daily_Spend": parse_decimal(row["Daily_Spend"], row_number, "Daily_Spend"),
                "Daily_Impressions": parse_decimal(
                    row["Daily_Impressions"],
                    row_number,
                    "Daily_Impressions",
                ),
                "Spend_Type": row["Spend_Type"],
            }
    return daily


def load_benchmark_daily() -> dict[str, dict[str, str | Decimal]]:
    daily: dict[str, dict[str, str | Decimal]] = {}
    with FINAL_ATMOSPHERE_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        validate_required_columns(
            reader.fieldnames,
            [
                "Date",
                "Vendor",
                "Brand",
                "Channel",
                "Platform",
                "Spend",
                "Impressions",
                "Data_Grain",
                "Source_File",
            ],
            FINAL_ATMOSPHERE_PATH,
        )
        for row_number, row in enumerate(reader, start=2):
            if SOURCE_TOKEN not in (row.get("Source_File") or ""):
                continue
            daily[row["Date"]] = {
                "Date": row["Date"],
                "Vendor": row["Vendor"],
                "Brand": row["Brand"],
                "Channel": row["Channel"],
                "Platform": row["Platform"],
                "Spend": parse_decimal(row["Spend"], row_number, "Spend"),
                "Impressions": parse_decimal(row["Impressions"], row_number, "Impressions"),
                "Data_Grain": row["Data_Grain"],
                "Source_File": row["Source_File"],
            }

    if not daily:
        raise ValueError(f"No benchmark rows containing {SOURCE_TOKEN!r} were found.")
    return daily


def summarize_daily(daily: dict[str, dict[str, str | Decimal]], spend_key: str, imps_key: str) -> dict[str, object]:
    dates = sorted(daily)
    return {
        "rows": len(daily),
        "min_date": dates[0] if dates else "",
        "max_date": dates[-1] if dates else "",
        "spend": sum((row[spend_key] for row in daily.values()), Decimal("0")),
        "impressions": sum((row[imps_key] for row in daily.values()), Decimal("0")),
    }


def compare_daily(
    generated: dict[str, dict[str, str | Decimal]],
    benchmark: dict[str, dict[str, str | Decimal]],
) -> dict[str, object]:
    generated_dates = set(generated)
    benchmark_dates = set(benchmark)
    shared_dates = sorted(generated_dates & benchmark_dates)

    column_mappings = [
        ("Date", "Date", "Date"),
        ("Vendor", "Vendor", "Vendor"),
        ("Brand", "Brand", "Brand"),
        ("Marketing_Channel", "Channel", "Channel"),
        ("Sub_Channel", "Platform", "Platform"),
        ("Daily_Spend", "Spend", "Spend"),
        ("Daily_Impressions", "Impressions", "Impressions"),
    ]

    mismatch_counts = {label: 0 for _, _, label in column_mappings}
    differing_rows = []
    max_spend_diff = Decimal("0")
    max_impression_diff = Decimal("0")

    for current_date in shared_dates:
        generated_row = generated[current_date]
        benchmark_row = benchmark[current_date]
        row_diffs: dict[str, tuple[object, object]] = {}

        for generated_column, benchmark_column, label in column_mappings:
            generated_value = generated_row[generated_column]
            benchmark_value = benchmark_row[benchmark_column]

            if label == "Spend":
                spend_diff = generated_value - benchmark_value
                max_spend_diff = max(max_spend_diff, abs(spend_diff))
                matched = abs(spend_diff) <= SPEND_TOLERANCE
            elif label == "Impressions":
                impression_diff = generated_value - benchmark_value
                max_impression_diff = max(max_impression_diff, abs(impression_diff))
                matched = abs(impression_diff) <= IMPRESSION_TOLERANCE
            else:
                matched = generated_value == benchmark_value

            if not matched:
                mismatch_counts[label] += 1
                row_diffs[label] = (generated_value, benchmark_value)

        if row_diffs:
            differing_rows.append({"Date": current_date, "Diffs": row_diffs})

    return {
        "generated_summary": summarize_daily(generated, "Daily_Spend", "Daily_Impressions"),
        "benchmark_summary": summarize_daily(benchmark, "Spend", "Impressions"),
        "missing_dates": sorted(benchmark_dates - generated_dates),
        "extra_dates": sorted(generated_dates - benchmark_dates),
        "mismatch_counts": mismatch_counts,
        "max_spend_diff": max_spend_diff,
        "max_impression_diff": max_impression_diff,
        "differing_rows": differing_rows,
        "generated_campaigns": sorted({row["Campaign"] for row in generated.values()}),
        "generated_spend_types": sorted({row["Spend_Type"] for row in generated.values()}),
        "benchmark_data_grains": sorted({row["Data_Grain"] for row in benchmark.values()}),
        "benchmark_source_files": sorted({row["Source_File"] for row in benchmark.values()}),
    }


def write_report(comparison: dict[str, object]) -> None:
    generated = comparison["generated_summary"]
    benchmark = comparison["benchmark_summary"]
    missing_dates = comparison["missing_dates"]
    extra_dates = comparison["extra_dates"]
    mismatch_counts = comparison["mismatch_counts"]
    differing_rows = comparison["differing_rows"]

    lines = [
        "# Atmosphere Comparison Report",
        "",
        "## Summary",
        "",
        "| Metric | Generated Atmosphere.csv | Final Atmosphere.csv |",
        "|---|---:|---:|",
        f"| Rows | {generated['rows']} | {benchmark['rows']} |",
        f"| Date range | {generated['min_date']} to {generated['max_date']} | "
        f"{benchmark['min_date']} to {benchmark['max_date']} |",
        f"| Total spend | {decimal_text(generated['spend'], '0.01')} | "
        f"{decimal_text(benchmark['spend'], '0.01')} |",
        f"| Total impressions | {decimal_text(generated['impressions'])} | "
        f"{decimal_text(benchmark['impressions'])} |",
        "",
        "## Date Coverage",
        "",
        f"- Missing dates from generated output: {len(missing_dates)}",
        f"- Extra dates in generated output: {len(extra_dates)}",
        "",
        "## Comparable Column Checks",
        "",
        "| Generated column | Final column | Mismatched rows |",
        "|---|---|---:|",
        f"| Date | Date | {mismatch_counts['Date']} |",
        f"| Vendor | Vendor | {mismatch_counts['Vendor']} |",
        f"| Brand | Brand | {mismatch_counts['Brand']} |",
        f"| Marketing_Channel | Channel | {mismatch_counts['Channel']} |",
        f"| Sub_Channel | Platform | {mismatch_counts['Platform']} |",
        f"| Daily_Spend | Spend | {mismatch_counts['Spend']} |",
        f"| Daily_Impressions | Impressions | {mismatch_counts['Impressions']} |",
        "",
        "## Difference Checks",
        "",
        f"- Max absolute daily spend difference: {comparison['max_spend_diff']}",
        f"- Max absolute daily impression difference: {comparison['max_impression_diff']}",
        f"- Material differing dates: {len(differing_rows)}",
        "",
        "## Schema-Specific Columns",
        "",
        f"- Generated `Campaign`: {', '.join(comparison['generated_campaigns'])}",
        f"- Generated `Spend_Type`: {', '.join(comparison['generated_spend_types'])}",
        f"- Final `Data_Grain`: {', '.join(comparison['benchmark_data_grains'])}",
        f"- Final `Source_File`: {', '.join(comparison['benchmark_source_files'])}",
        "",
    ]

    if missing_dates:
        lines.extend(["### Missing Dates", "", ", ".join(missing_dates[:25]), ""])
    if extra_dates:
        lines.extend(["### Extra Dates", "", ", ".join(extra_dates[:25]), ""])

    if differing_rows:
        lines.extend(["### Sample Differing Dates", ""])
        for row in differing_rows[:10]:
            lines.append(f"- {row['Date']}: {row['Diffs']}")
        lines.append("")
    else:
        lines.extend(
            [
                "No material differences were found across comparable columns.",
                "",
            ]
        )

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    rows = build_daily_rows()
    write_csv(rows)

    generated = load_generated_daily()
    benchmark = load_benchmark_daily()
    comparison = compare_daily(generated, benchmark)
    write_report(comparison)

    print(f"Wrote {len(rows)} rows to {OUTPUT_PATH}")
    print(f"Wrote comparison report to {REPORT_PATH}")


if __name__ == "__main__":
    main()
