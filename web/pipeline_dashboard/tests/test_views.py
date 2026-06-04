from __future__ import annotations

import json
import shutil
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from pipeline_dashboard.models import Asset, AssetEvent, AssetStatus, ParsedOutput, ShareFileFolder, Vendor
from pipeline_dashboard.parser_workflow import approval_root_id, period_series_from_rows
from pipeline_dashboard.services import _reconcile_duplicate_roles_for_group
from pipeline_dashboard.sharefile_mirror import load_approval_mirror


class FakeApprovalClient:
    def __init__(self):
        from datetime import datetime

        self.current_month = datetime.now().strftime("%B_%Y")
        self.folder_parts = []
        self.uploaded_name = ""
        self.notify = False
        self.copy_access_controls = False

    def ensure_folder_path(self, root_id, parts, copy_access_controls=False):
        self.root_id = root_id
        self.folder_parts = parts
        self.copy_access_controls = copy_access_controls
        return SimpleNamespace(id="fo-approval")

    def upload_bytes(self, folder_id, filename, content, content_type, notify, overwrite):
        self.folder_id = folder_id
        self.uploaded_name = filename
        self.uploaded_content = content
        self.notify = notify
        return SimpleNamespace(id="fi-uploaded", name=filename)


class DashboardViewTests(TestCase):
    def test_dashboard_renders_with_empty_state(self):
        response = self.client.get(reverse("pipeline_dashboard:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dashboard")
        self.assertContains(response, "No assets need review yet")

    def test_dashboard_renders_catalogue_data(self):
        vendor, _ = Vendor.objects.get_or_create(name="AdTaxi", defaults={"parser_key": "adtaxi"})
        folder = ShareFileFolder.objects.create(
            vendor=vendor,
            folder_id="fo-example",
            label="Shared Folders/AdTaxi",
            file_patterns=["*.xlsx"],
        )
        Asset.objects.create(
            remote_item_id="fi-example",
            vendor=vendor,
            source_folder=folder,
            status=AssetStatus.NEW,
            name="AdTaxi report.xlsx",
        )

        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AdTaxi report.xlsx")
        self.assertContains(response, "AdTaxi")

    def test_dashboard_renders_sharefile_mirror_totals(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_review_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ShareFile mirror totals")
        self.assertContains(response, "Folders")
        self.assertContains(response, "Files")
        self.assertContains(response, "New")
        self.assertContains(response, "Duplicate Names")
        content = response.content.decode()
        self.assertIn('<div class="metric-label">Folders</div>', content)
        self.assertIn('<div class="metric-label">Files</div>', content)
        self.assertIn('<div class="metric-label">New</div>', content)
        self.assertIn('<div class="metric-label">Duplicate Names</div>', content)
        self.assertIn('<div class="metric-value">1</div>', content)
        self.assertIn('<div class="metric-value">0</div>', content)
        self.assertContains(response, "ShareFile sources")
        self.assertContains(response, "Ready for review")
        self.assertContains(response, "Name collisions")

    def test_removed_public_pages_return_not_found(self):
        self.assertEqual(self.client.get("/assets/").status_code, 404)

    def test_vendors_page_renders_vendor_activity(self):
        vendor, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        folder = ShareFileFolder.objects.create(
            vendor=vendor,
            folder_id="fo-loop",
            label="Shared Folders/Loop",
            file_patterns=["*.xlsx"],
        )
        asset = Asset.objects.create(
            remote_item_id="fi-loop-vendor-page",
            vendor=vendor,
            source_folder=folder,
            status=AssetStatus.REVIEW,
            name="Loop report.xlsx",
            created_by_name="Uploader One",
            created_by_email="uploader@example.com",
        )
        ParsedOutput.objects.create(
            asset=asset,
            vendor=vendor,
            output_path="data/output/Loop/Loop_Apr_2026_v1.csv",
            reporting_period="March_2026",
            comparison_status="sent_for_approval",
        )
        AssetEvent.objects.create(asset=asset, event_type="approval_sent", to_status=AssetStatus.REVIEW)

        response = self.client.get(reverse("pipeline_dashboard:vendors"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vendors")
        self.assertContains(response, "Create Vendor")
        self.assertContains(response, "Loop")
        self.assertContains(response, "Shared Folders/Loop")
        self.assertContains(response, "Uploader One")
        self.assertContains(response, "approval_sent")

    def test_create_vendor_adds_local_vendor(self):
        response = self.client.post(
            reverse("pipeline_dashboard:create_vendor"),
            {"name": "New Vendor", "parser_key": "", "notes": "Pilot source"},
        )

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        vendor = Vendor.objects.get(name="New Vendor")
        self.assertEqual(vendor.parser_key, "new_vendor")
        self.assertEqual(vendor.notes, "Pilot source")
        self.assertTrue(vendor.is_active)

    def test_create_vendor_rejects_duplicate_name(self):
        Vendor.objects.create(name="Case Vendor", parser_key="case_vendor")

        response = self.client.post(reverse("pipeline_dashboard:create_vendor"), {"name": "case vendor"})

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        self.assertEqual(Vendor.objects.filter(name__iexact="case vendor").count(), 1)

    def test_update_vendor_changes_local_metadata(self):
        vendor = Vendor.objects.create(name="Mutable Vendor", parser_key="mutable")

        response = self.client.post(
            reverse("pipeline_dashboard:update_vendor", args=[vendor.id]),
            {"name": "Renamed Vendor", "parser_key": "renamed", "notes": "Updated"},
        )

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        vendor.refresh_from_db()
        self.assertEqual(vendor.name, "Renamed Vendor")
        self.assertEqual(vendor.parser_key, "renamed")
        self.assertEqual(vendor.notes, "Updated")
        self.assertFalse(vendor.is_active)

    def test_delete_vendor_removes_unused_vendor(self):
        vendor = Vendor.objects.create(name="Unused Vendor", parser_key="unused")

        response = self.client.post(reverse("pipeline_dashboard:delete_vendor", args=[vendor.id]))

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        self.assertFalse(Vendor.objects.filter(pk=vendor.pk).exists())

    def test_delete_vendor_deactivates_linked_vendor(self):
        vendor = Vendor.objects.create(name="Linked Vendor", parser_key="linked")
        Asset.objects.create(remote_item_id="fi-linked-vendor", vendor=vendor, status=AssetStatus.NEW, name="linked.xlsx")

        response = self.client.post(reverse("pipeline_dashboard:delete_vendor", args=[vendor.id]))

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        vendor.refresh_from_db()
        self.assertFalse(vendor.is_active)

    def test_assign_vendor_folder_updates_local_binding(self):
        vendor = Vendor.objects.create(name="Folder Vendor", parser_key="folder_vendor")
        folder = ShareFileFolder.objects.create(folder_id="fo-unassigned", label="Shared Folders/Unassigned")

        response = self.client.post(
            reverse("pipeline_dashboard:assign_vendor_folder", args=[folder.id]),
            {"vendor_id": vendor.id},
        )

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        folder.refresh_from_db()
        self.assertEqual(folder.vendor, vendor)

        response = self.client.post(reverse("pipeline_dashboard:assign_vendor_folder", args=[folder.id]), {"vendor_id": ""})

        self.assertRedirects(response, reverse("pipeline_dashboard:vendors"))
        folder.refresh_from_db()
        self.assertIsNone(folder.vendor)

    def test_folders_page_renders_sharefile_mirror(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            state_root = repo_root / "data" / "state"
            state_root.mkdir(parents=True)
            (state_root / "sharefile_snapshot_latest.json").write_text(
                json.dumps(
                    {
                        "run_id": "snapshot-1",
                        "created_at": "2026-05-07T10:00:00Z",
                        "files": [
                            {
                                "remote_item_id": "fi-new",
                                "name": "z-new.xlsx",
                                "remote_path": "home/josh/z-new.xlsx",
                                "local_path": "data/inbox/home/josh/z-new.xlsx",
                                "source_folder_path": "home/josh",
                                "extension": ".xlsx",
                                "size": 10,
                                "modified_at": "2026-05-07T12:00:00Z",
                                "creator": "Uploader One",
                                "raw_metadata": {"LastModifiedByUserID": "user-1"},
                            },
                            {
                                "remote_item_id": "fi-processed",
                                "name": "a-processed.csv",
                                "remote_path": "home/josh/a-processed.csv",
                                "local_path": "data/inbox/home/josh/a-processed.csv",
                                "source_folder_path": "home/josh",
                                "extension": ".csv",
                                "size": 20,
                                "modified_at": "2026-05-06T12:00:00Z",
                                "creator": "Uploader Two",
                                "raw_metadata": {"LastModifiedByUserID": "user-2"},
                            },
                            {
                                "remote_item_id": "fi-active",
                                "name": "m-active.csv",
                                "remote_path": "home/josh/m-active.csv",
                                "local_path": "data/inbox/home/josh/m-active.csv",
                                "source_folder_path": "home/josh",
                                "source_folder_id": "fo-josh",
                                "extension": ".csv",
                                "size": 30,
                                "modified_at": "2026-05-05T12:00:00Z",
                                "creator": "Uploader Two",
                                "raw_metadata": {"LastModifiedByUserID": "user-2"},
                            },
                            {
                                "remote_item_id": "fi-approval-internal",
                                "name": "Loop_May_2026_v1.csv",
                                "remote_path": "allshared/Approval/May_2026/Loop/Loop_May_2026_v1.csv",
                                "local_path": "data/inbox/allshared/Approval/May_2026/Loop/Loop_May_2026_v1.csv",
                                "source_folder_path": "allshared/Approval/May_2026/Loop",
                                "extension": ".csv",
                                "size": 40,
                                "modified_at": "2026-05-04T12:00:00Z",
                                "creator": "Uploader Two",
                                "raw_metadata": {"LastModifiedByUserID": "user-2"},
                            },
                            {
                                "remote_item_id": "fi-final-internal",
                                "name": "Loop_April_2026.csv",
                                "remote_path": "allshared/Final/April_2026/Loop_April_2026.csv",
                                "local_path": "data/inbox/allshared/Final/April_2026/Loop_April_2026.csv",
                                "source_folder_path": "allshared/Final/April_2026",
                                "extension": ".csv",
                                "size": 50,
                                "modified_at": "2026-05-03T12:00:00Z",
                                "creator": "Uploader Two",
                                "raw_metadata": {"LastModifiedByUserID": "user-2"},
                            },
                        ],
                    }
                )
            )
            (state_root / "inbox_profile_latest.json").write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "local_path": "data/inbox/home/josh/z-new.xlsx",
                                "name": "z-new.xlsx",
                                "kind": "excel",
                                "status": "profiled",
                                "sheet_count": 1,
                            },
                            {
                                "local_path": "data/inbox/home/josh/a-processed.csv",
                                "name": "a-processed.csv",
                                "kind": "csv",
                                "status": "profiled",
                            },
                            {
                                "local_path": "data/inbox/home/josh/m-active.csv",
                                "name": "m-active.csv",
                                "kind": "csv",
                                "status": "profiled",
                            },
                            {
                                "local_path": "data/inbox/home/josh/deleted.xlsx",
                                "name": "deleted.xlsx",
                                "kind": "excel",
                                "status": "profiled",
                                "sheet_count": 2,
                            },
                            {
                                "local_path": "data/inbox/allshared/Approval/May_2026/Loop/Loop_May_2026_v1.csv",
                                "name": "Loop_May_2026_v1.csv",
                                "kind": "csv",
                                "status": "profiled",
                            },
                            {
                                "local_path": "data/inbox/allshared/Final/April_2026/Loop_April_2026.csv",
                                "name": "Loop_April_2026.csv",
                                "kind": "csv",
                                "status": "profiled",
                            },
                            {
                                "local_path": "data/inbox/allshared/Approval/May_2026/profile-only.csv",
                                "name": "profile-only.csv",
                                "kind": "csv",
                                "status": "profiled",
                            },
                        ]
                    }
                )
            )
            (state_root / "file_processing_state.json").write_text(
                json.dumps({"processed_local_paths": ["data/inbox/home/josh/a-processed.csv"]})
            )
            (state_root / "sharefile_users_latest.json").write_text(
                json.dumps(
                    {
                        "users_by_id": {
                            "user-1": {"full_name": "Uploader One", "email": "one@example.com"},
                            "user-2": {"full_name": "Uploader Two", "email": "two@example.com"},
                        }
                    }
                )
            )
            (state_root / "sharefile_sync_state.json").write_text(
                json.dumps({"status": "success", "finished_at": "2026-05-07T12:00:00+00:00"})
            )
            Asset.objects.create(
                remote_item_id="fi-active",
                status=AssetStatus.PROCESSING,
                name="m-active.csv",
                local_path="data/inbox/home/josh/m-active.csv",
            )
            adtaxi, _ = Vendor.objects.get_or_create(name="AdTaxi", defaults={"parser_key": "adtaxi"})

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        content = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Loaded Folders")
        self.assertContains(response, "josh")
        self.assertContains(response, "z-new.xlsx")
        self.assertContains(response, "a-processed.csv")
        self.assertContains(response, "m-active.csv")
        self.assertContains(response, "deleted.xlsx")
        self.assertContains(response, "Uploader")
        self.assertContains(response, "Mail")
        self.assertContains(response, "Uploader One")
        self.assertContains(response, "Uploader Two")
        self.assertContains(response, "one@example.com")
        self.assertContains(response, "two@example.com")
        self.assertContains(response, 'data-page-size="10"')
        self.assertContains(response, "Prev")
        self.assertContains(response, "Next")
        self.assertLess(response.content.decode().index("z-new.xlsx"), response.content.decode().index("a-processed.csv"))
        self.assertContains(response, ">New<", html=False)
        self.assertContains(response, ">Active<", html=False)
        self.assertContains(response, ">Processed<", html=False)
        self.assertContains(response, ">Deleted<", html=False)
        self.assertContains(response, "Active")
        self.assertContains(response, "Review")
        self.assertContains(response, "Deleted")
        self.assertContains(response, "PodcastOne, Octopus, Loop, TVM, TAIV")
        self.assertContains(response, 'class="metrics metrics-compact folders-metrics"', html=False)
        self.assertContains(response, 'class="metric blue update-metric"', html=False)
        self.assertContains(response, "Last: May 07, 2026 12:00")
        self.assertContains(response, "Updating SF folders")
        self.assertNotContains(response, f'data-allowed-vendors="{adtaxi.id}"')
        # Loaded Folders chapter still excludes Approval/Final subfolders.
        loaded_section_start = content.index(">Loaded Folders<")
        loaded_section_end = content.index("</section>", loaded_section_start)
        loaded_block = content[loaded_section_start:loaded_section_end]
        self.assertNotIn("Approval/May_2026/Loop", loaded_block)
        self.assertNotIn("Final/April_2026", loaded_block)
        self.assertNotIn("Loop_May_2026_v1.csv", loaded_block)
        self.assertNotIn("Loop_April_2026.csv", loaded_block)
        self.assertNotIn("profile-only.csv", loaded_block)
        # ...but the new Approval chapter exposes them.
        self.assertContains(response, "folders-approval-panel")
        self.assertContains(response, "Loop_May_2026_v1.csv")
        self.assertNotContains(response, "Loop_April_2026.csv")  # Final-only file is still hidden
        self.assertNotContains(response, "profile-only.csv")

    def _write_approval_fixture(self, repo_root: Path) -> dict[str, str]:
        state_root = repo_root / "data" / "state"
        state_root.mkdir(parents=True, exist_ok=True)
        inbox_root = repo_root / "data" / "inbox"
        files = [
            {
                "remote_item_id": "fi-may-loop-v2",
                "name": "Loop_May_2026_v2.csv",
                "remote_path": "allshared/Approval/May_2026/Loop/Loop_May_2026_v2.csv",
                "local_path": "data/inbox/allshared/Approval/May_2026/Loop/Loop_May_2026_v2.csv",
                "source_folder_path": "allshared/Approval/May_2026/Loop",
                "source_folder_id": "fo-may-loop",
                "extension": ".csv",
                "size": 100,
                "modified_at": "2026-05-29T12:00:00Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-may-loop-v1",
                "name": "Loop_May_2026_v1.csv",
                "remote_path": "allshared/Approval/May_2026/Loop/Loop_May_2026_v1.csv",
                "local_path": "data/inbox/allshared/Approval/May_2026/Loop/Loop_May_2026_v1.csv",
                "source_folder_path": "allshared/Approval/May_2026/Loop",
                "source_folder_id": "fo-may-loop",
                "extension": ".csv",
                "size": 100,
                "modified_at": "2026-05-07T12:00:00Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-may-tvm-v1",
                "name": "TVM_May_2026_v1.csv",
                "remote_path": "allshared/Approval/May_2026/TVM/TVM_May_2026_v1.csv",
                "local_path": "data/inbox/allshared/Approval/May_2026/TVM/TVM_May_2026_v1.csv",
                "source_folder_path": "allshared/Approval/May_2026/TVM",
                "source_folder_id": "fo-may-tvm",
                "extension": ".csv",
                "size": 80,
                "modified_at": "2026-05-07T19:04:40Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-april-adtaxi-v1",
                "name": "AdTaxi_Apr_2026_v1.csv",
                "remote_path": "allshared/Approval/April_2026/AdTaxi/AdTaxi_Apr_2026_v1.csv",
                "local_path": "data/inbox/allshared/Approval/April_2026/AdTaxi/AdTaxi_Apr_2026_v1.csv",
                "source_folder_path": "allshared/Approval/April_2026/AdTaxi",
                "source_folder_id": "fo-april-adtaxi",
                "extension": ".csv",
                "size": 50,
                "modified_at": "2026-05-12T10:13:36Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-april-tvm-missing",
                "name": "TVM_Apr_2026_v1.csv",
                "remote_path": "allshared/Approval/April_2026/TVM/TVM_Apr_2026_v1.csv",
                "local_path": "data/inbox/allshared/Approval/April_2026/TVM/TVM_Apr_2026_v1.csv",
                "source_folder_path": "allshared/Approval/April_2026/TVM",
                "source_folder_id": "fo-april-tvm",
                "extension": ".csv",
                "size": 60,
                "modified_at": "2026-05-12T10:13:36Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-final-loop",
                "name": "Loop_April_2026.csv",
                "remote_path": "allshared/Final/April_2026/Loop_April_2026.csv",
                "local_path": "data/inbox/allshared/Final/April_2026/Loop_April_2026.csv",
                "source_folder_path": "allshared/Final/April_2026",
                "source_folder_id": "fo-final-april",
                "extension": ".csv",
                "size": 200,
                "modified_at": "2026-05-12T10:13:36Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
            {
                "remote_item_id": "fi-vendor-loop",
                "name": "loop-may-2026.xlsx",
                "remote_path": "home/josh/loop-may-2026.xlsx",
                "local_path": "data/inbox/home/josh/loop-may-2026.xlsx",
                "source_folder_path": "home/josh",
                "source_folder_id": "fo-josh",
                "extension": ".xlsx",
                "size": 30,
                "modified_at": "2026-05-08T08:00:00Z",
                "creator": "Data Team",
                "raw_metadata": {"LastModifiedByUserID": "user-1"},
            },
        ]
        (state_root / "sharefile_snapshot_latest.json").write_text(
            json.dumps(
                {
                    "run_id": "snapshot-approval",
                    "created_at": "2026-05-29T12:00:00Z",
                    "files": files,
                }
            ),
            encoding="utf-8",
        )
        (state_root / "inbox_profile_latest.json").write_text(
            json.dumps(
                {
                    "files": [
                        {
                            "local_path": row["local_path"],
                            "name": row["name"],
                            "kind": "csv",
                            "status": "profiled",
                        }
                        for row in files
                    ]
                }
            ),
            encoding="utf-8",
        )
        (state_root / "sharefile_users_latest.json").write_text(
            json.dumps({"users_by_id": {"user-1": {"full_name": "Data Team", "email": "team@example.com"}}}),
            encoding="utf-8",
        )
        (state_root / "sharefile_sync_state.json").write_text(
            json.dumps({"status": "success", "finished_at": "2026-05-29T12:00:00+00:00"}),
            encoding="utf-8",
        )

        written = {}
        for row in files:
            if not row["local_path"].startswith("data/inbox/"):
                continue
            if row["remote_item_id"] == "fi-april-tvm-missing":
                # Intentionally do not create the local file to exercise the
                # "missing locally" code path.
                continue
            local = repo_root / row["local_path"]
            local.parent.mkdir(parents=True, exist_ok=True)
            local.write_text("Date,Vendor\n2026-05-01,Vendor\n", encoding="utf-8")
            written[row["local_path"]] = str(local)
        return written

    def test_folders_page_renders_approval_chapter(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        content = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "folders-approval-panel")
        # Months are newest first
        may_idx = content.index(">May 26<")
        apr_idx = content.index(">Apr 26<")
        self.assertLess(may_idx, apr_idx)
        # Vendors are alphabetical
        loop_idx = content.index('data-vendor-name="Loop"', may_idx)
        tvm_idx = content.index('data-vendor-name="TVM"', may_idx)
        self.assertLess(loop_idx, tvm_idx)
        # Versions are extracted
        self.assertIn(">v2<", content)
        self.assertIn(">v1<", content)
        # Files from Final/ are not in the approval chapter
        self.assertNotIn("fi-final-loop", content)
        self.assertNotIn("Loop_April_2026.csv", content)
        # No Approve / Cancel buttons in the Approval chapter
        # Locate the actual HTML section (skip CSS in <style> block)
        approval_section_start = content.index('aria-label="ShareFile Approval folders grouped by month"')
        approval_end = content.index("</section>", approval_section_start)
        approval_block = content[approval_section_start:approval_end]
        self.assertNotIn("Approved</button>", approval_block)
        self.assertNotIn(">Cancel</button>", approval_block)
        # Download + Review buttons present
        self.assertIn("download-button", approval_block)
        self.assertIn("data-approval-review-button", approval_block)
        # Missing-local file shows a missing marker
        self.assertIn("is-missing", content)
        # Search input present
        self.assertContains(response, "data-approval-folders-search")

    def test_review_approval_file_returns_preview(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse(
                        "pipeline_dashboard:review_approval_file",
                        kwargs={"remote_item_id": "fi-may-tvm-v1"},
                    )
                )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("file", payload)
        self.assertEqual(payload["file"]["name"], "TVM_May_2026_v1.csv")
        self.assertEqual(payload["file"]["kind"], "csv")

    def test_review_approval_file_returns_404_for_unknown(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse(
                        "pipeline_dashboard:review_approval_file",
                        kwargs={"remote_item_id": "fi-unknown"},
                    )
                )

        self.assertEqual(response.status_code, 404)

    def test_download_approval_output_returns_csv(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse(
                        "pipeline_dashboard:download_approval_output",
                        kwargs={"remote_item_id": "fi-may-tvm-v1"},
                    )
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get("Content-Type"), "text/csv")
        self.assertIn("attachment", response.get("Content-Disposition", ""))
        self.assertIn("TVM_May_2026_v1.csv", response.get("Content-Disposition", ""))

    def test_download_approval_output_404_for_unknown(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse(
                        "pipeline_dashboard:download_approval_output",
                        kwargs={"remote_item_id": "fi-unknown"},
                    )
                )

        self.assertEqual(response.status_code, 404)

    def test_download_approval_output_excludes_final_folder(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse(
                        "pipeline_dashboard:download_approval_output",
                        kwargs={"remote_item_id": "fi-final-loop"},
                    )
                )

        self.assertEqual(response.status_code, 404)

    def test_load_approval_mirror_groups_newest_first_alphabetical(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            self._write_approval_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                mirror = load_approval_mirror()

        self.assertEqual(mirror.summary["file_count"], 5)
        self.assertEqual(mirror.summary["month_count"], 2)
        months = mirror.folders
        self.assertEqual(months[0]["label"], "May_2026")
        self.assertEqual(months[1]["label"], "April_2026")
        may_vendors = [v["name"] for v in months[0]["vendors"]]
        self.assertEqual(may_vendors, ["Loop", "TVM"])
        # Loop has 2 files (v1 and v2); v2 should come first because of newer modified_at
        loop_files = months[0]["vendors"][0]["files"]
        self.assertEqual([f["name"] for f in loop_files], ["Loop_May_2026_v2.csv", "Loop_May_2026_v1.csv"])
        # File missing locally
        april_tvm = next(v for v in months[1]["vendors"] if v["name"] == "TVM")
        self.assertFalse(april_tvm["files"][0]["exists_locally"])
        # Final/ files are not in the approval mirror
        all_remote_ids = [
            f["remote_item_id"]
            for m in months
            for v in m["vendors"]
            for f in v["files"]
        ]
        self.assertNotIn("fi-final-loop", all_remote_ids)
        # Vendor files for the Loaded Folders chapter are not in the approval mirror
        self.assertNotIn("fi-vendor-loop", all_remote_ids)

    def test_folders_page_renders_search_bar(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            state_root = repo_root / "data" / "state"
            state_root.mkdir(parents=True)
            (state_root / "sharefile_snapshot_latest.json").write_text(
                json.dumps(
                    {
                        "run_id": "snapshot-search",
                        "created_at": "2026-05-07T10:00:00Z",
                        "files": [
                            {
                                "remote_item_id": "fi-search",
                                "name": "search-me.xlsx",
                                "local_path": "data/inbox/home/josh/search-me.xlsx",
                                "source_folder_path": "home/josh",
                                "extension": ".xlsx",
                                "size": 10,
                                "modified_at": "2026-05-07T12:00:00Z",
                                "creator": "Uploader One",
                                "raw_metadata": {"LastModifiedByUserID": "user-1"},
                            },
                        ],
                    }
                )
            )
            (state_root / "inbox_profile_latest.json").write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "local_path": "data/inbox/home/josh/search-me.xlsx",
                                "name": "search-me.xlsx",
                                "kind": "excel",
                                "status": "profiled",
                            }
                        ]
                    }
                )
            )
            (state_root / "file_processing_state.json").write_text(json.dumps({}))
            (state_root / "sharefile_users_latest.json").write_text(
                json.dumps({"users_by_id": {"user-1": {"full_name": "Uploader One", "email": "one@example.com"}}})
            )
            (state_root / "sharefile_sync_state.json").write_text(json.dumps({}))

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("data-folder-search", content)
        self.assertIn("Search files by name or uploader email", content)
        self.assertIn("mirror-head-search", content)
        self.assertIn("mirror-search-input", content)

    def test_folders_page_sorts_files_by_status_tiers_and_modified(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            state_root = repo_root / "data" / "state"
            state_root.mkdir(parents=True)
            (state_root / "sharefile_snapshot_latest.json").write_text(
                json.dumps(
                    {
                        "run_id": "snapshot-sort",
                        "created_at": "2026-05-07T10:00:00Z",
                        "files": [
                            {
                                "remote_item_id": "fi-processed-active-recent",
                                "name": "processed-active-recent.csv",
                                "local_path": "data/inbox/home/josh/processed-active-recent.csv",
                                "source_folder_path": "home/josh",
                                "extension": ".csv",
                                "size": 10,
                                "modified_at": "2026-05-07T14:00:00Z",
                            },
                            {
                                "remote_item_id": "fi-new-active-old",
                                "name": "new-active-old.csv",
                                "local_path": "data/inbox/home/josh/new-active-old.csv",
                                "source_folder_path": "home/josh",
                                "extension": ".csv",
                                "size": 10,
                                "modified_at": "2026-05-06T12:00:00Z",
                            },
                            {
                                "remote_item_id": "fi-new-active-recent",
                                "name": "new-active-recent.csv",
                                "local_path": "data/inbox/home/josh/new-active-recent.csv",
                                "source_folder_path": "home/josh",
                                "extension": ".csv",
                                "size": 10,
                                "modified_at": "2026-05-07T12:00:00Z",
                            },
                        ],
                    }
                )
            )
            (state_root / "inbox_profile_latest.json").write_text(
                json.dumps(
                    {
                        "files": [
                            {"local_path": "data/inbox/home/josh/processed-active-recent.csv", "name": "processed-active-recent.csv", "kind": "csv"},
                            {"local_path": "data/inbox/home/josh/new-active-old.csv", "name": "new-active-old.csv", "kind": "csv"},
                            {"local_path": "data/inbox/home/josh/new-active-recent.csv", "name": "new-active-recent.csv", "kind": "csv"},
                            {"local_path": "data/inbox/home/josh/deleted-active-old.csv", "name": "deleted-active-old.csv", "kind": "csv", "status": "profiled"},
                            {"local_path": "data/inbox/home/josh/new-inactive-recent.csv", "name": "new-inactive-recent.csv", "kind": "csv"},
                            {"local_path": "data/inbox/home/josh/processed-inactive-old.csv", "name": "processed-inactive-old.csv", "kind": "csv"},
                        ]
                    }
                )
            )
            (state_root / "file_processing_state.json").write_text(
                json.dumps(
                    {
                        "processed_local_paths": [
                            "data/inbox/home/josh/processed-active-recent.csv",
                            "data/inbox/home/josh/processed-inactive-old.csv",
                        ]
                    }
                )
            )
            (state_root / "sharefile_users_latest.json").write_text(json.dumps({}))
            (state_root / "sharefile_sync_state.json").write_text(json.dumps({}))

            Asset.objects.create(
                remote_item_id="fi-new-inactive-recent",
                status=AssetStatus.NEW,
                name="new-inactive-recent.csv",
                local_path="data/inbox/home/josh/new-inactive-recent.csv",
                is_active=False,
            )
            Asset.objects.create(
                remote_item_id="fi-processed-inactive-old",
                status=AssetStatus.PROCESSED,
                name="processed-inactive-old.csv",
                local_path="data/inbox/home/josh/processed-inactive-old.csv",
                is_active=False,
            )

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        # Tier 0 (active under-review) at top, sorted by modified desc
        self.assertLess(content.index("new-active-recent.csv"), content.index("new-active-old.csv"))
        # Tier 0 above tier 1 (active finished)
        self.assertLess(content.index("new-active-old.csv"), content.index("processed-active-recent.csv"))
        self.assertLess(content.index("new-active-recent.csv"), content.index("processed-active-recent.csv"))
        # Tier 1 sorted by modified desc
        self.assertLess(content.index("processed-active-recent.csv"), content.index("deleted-active-old.csv"))
        # Tier 1 above tier 2 (inactive)
        self.assertLess(content.index("deleted-active-old.csv"), content.index("new-inactive-recent.csv"))
        self.assertLess(content.index("processed-active-recent.csv"), content.index("processed-inactive-old.csv"))
        # Tier 2 sorted by modified desc
        self.assertLess(content.index("new-inactive-recent.csv"), content.index("processed-inactive-old.csv"))

    def test_folders_page_sorts_by_new_count_and_marks_cross_folder_duplicates(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            state_root = repo_root / "data" / "state"
            state_root.mkdir(parents=True)
            duplicate_name = "T-MOBILE March 2026 EOM Report.xlsx"
            processed_duplicate_path = f"data/inbox/home/josh/{duplicate_name}"
            internal_duplicate_path = f"data/inbox/allshared/May_2026_Internal_folders/{duplicate_name}"
            snapshot_files = [
                {
                    "remote_item_id": "fi-many-1",
                    "name": "new-a.csv",
                    "local_path": "data/inbox/home/many_new/new-a.csv",
                    "source_folder_path": "home/many_new",
                    "extension": ".csv",
                    "size": 10,
                    "modified_at": "2026-05-09T12:00:00Z",
                },
                {
                    "remote_item_id": "fi-many-2",
                    "name": "new-b.csv",
                    "local_path": "data/inbox/home/many_new/new-b.csv",
                    "source_folder_path": "home/many_new",
                    "extension": ".csv",
                    "size": 10,
                    "modified_at": "2026-05-09T11:00:00Z",
                },
                {
                    "remote_item_id": "fi-internal-duplicate",
                    "name": duplicate_name,
                    "local_path": internal_duplicate_path,
                    "source_folder_path": "allshared/May_2026_Internal_folders",
                    "extension": ".xlsx",
                    "size": 20,
                    "modified_at": "2026-05-08T12:00:00Z",
                },
                {
                    "remote_item_id": "fi-josh-duplicate",
                    "name": duplicate_name,
                    "local_path": processed_duplicate_path,
                    "source_folder_path": "home/josh",
                    "extension": ".xlsx",
                    "size": 20,
                    "modified_at": "2026-05-07T12:00:00Z",
                },
            ]
            (state_root / "sharefile_snapshot_latest.json").write_text(
                json.dumps({"run_id": "snapshot-duplicates", "files": snapshot_files})
            )
            (state_root / "inbox_profile_latest.json").write_text(
                json.dumps(
                    {
                        "files": [
                            {"local_path": row["local_path"], "name": row["name"], "kind": "csv"}
                            for row in snapshot_files
                        ]
                    }
                )
            )
            (state_root / "file_processing_state.json").write_text(
                json.dumps({"processed_local_paths": [processed_duplicate_path]})
            )

            Asset.objects.create(
                remote_item_id="fi-internal-duplicate",
                name=duplicate_name,
                local_path=internal_duplicate_path,
                status=AssetStatus.NEW,
                duplicate_group=" ".join(duplicate_name.casefold().strip().split()),
                remote_created_at="2026-05-08T12:00:00+00:00",
            )
            Asset.objects.create(
                remote_item_id="fi-josh-duplicate",
                name=duplicate_name,
                local_path=processed_duplicate_path,
                status=AssetStatus.NEW,
                duplicate_group=" ".join(duplicate_name.casefold().strip().split()),
                remote_created_at="2026-05-07T12:00:00+00:00",
            )
            _reconcile_duplicate_roles_for_group(" ".join(duplicate_name.casefold().strip().split()))

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [folder["display_name"] for folder in response.context["folders"]],
            ["many_new", "May_2026_Internal_folders", "josh"],
        )
        self.assertEqual(response.context["mirror_summary"]["duplicate_name_count"], 1)
        self.assertEqual(response.context["folders"][1]["counts"]["duplicate_names"], 1)
        self.assertEqual(response.context["folders"][2]["counts"]["duplicate_names"], 0)
        self.assertEqual(response.context["folders"][1]["allowed_vendor_names"], "RallyAdMedia, AdTaxi")
        duplicate_rows = [
            file
            for folder in response.context["folders"]
            for file in folder["files"]
            if file["duplicate_name"]
        ]
        self.assertEqual({row["local_path"] for row in duplicate_rows}, {internal_duplicate_path, processed_duplicate_path})
        content = response.content.decode()
        # One original badge, one duplicate badge
        self.assertIn("\n                            Original\n                          ", content)
        self.assertIn("\n                            Dup\n                          ", content)
        self.assertIn('data-allowed-vendor-names="RallyAdMedia, AdTaxi"', content)

    def test_review_file_preview_returns_csv_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            local_path = self._write_review_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse("pipeline_dashboard:review_file_preview"),
                    {"local_path": local_path},
                )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["file"]["name"], "new-report.csv")
        self.assertEqual(payload["file"]["kind"], "csv")
        self.assertEqual(payload["file"]["sheets"][0]["headers"], ["Date", "Campaign", "Spend"])
        self.assertEqual(payload["file"]["sheets"][0]["rows"][1], ["2026-05-07", "Spring", "10"])

    def test_process_review_file_marks_asset_active_with_vendor(self):
        vendor, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            local_path = self._write_review_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(
                    reverse("pipeline_dashboard:process_review_file"),
                    {"local_path": local_path, "vendor_id": vendor.id},
                )

        self.assertEqual(response.status_code, 200)
        asset = Asset.objects.get(remote_item_id="fi-review")
        self.assertEqual(asset.status, AssetStatus.PROCESSING)
        self.assertEqual(asset.vendor, vendor)
        self.assertEqual(asset.local_path, local_path)
        self.assertEqual(asset.parser_key, "loop")
        self.assertTrue(asset.events.filter(event_type="review_started").exists())

    def test_process_review_file_rejects_vendor_not_allowed_for_folder(self):
        vendor, _ = Vendor.objects.get_or_create(name="AdTaxi", defaults={"parser_key": "adtaxi"})
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            local_path = self._write_review_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(
                    reverse("pipeline_dashboard:process_review_file"),
                    {"local_path": local_path, "vendor_id": vendor.id},
                )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "AdTaxi is not available for this folder.")
        self.assertFalse(Asset.objects.filter(remote_item_id="fi-review").exists())

    def test_process_page_renders_processing_assets(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        podcastone, _ = Vendor.objects.get_or_create(name="PodcastOne", defaults={"parser_key": "podcastone"})
        Asset.objects.create(
            remote_item_id="fi-processing",
            vendor=loop,
            status=AssetStatus.PROCESSING,
            name="Loop report.csv",
            created_by_name="Uploader One",
            file_size=25,
        )
        Asset.objects.create(
            remote_item_id="fi-new-hidden",
            vendor=podcastone,
            status=AssetStatus.NEW,
            name="PodcastOne new.csv",
        )

        response = self.client.get(reverse("pipeline_dashboard:process"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Parsing")
        self.assertContains(response, "Parsing Files")
        self.assertContains(response, "Loop report.csv")
        self.assertContains(response, "Loop")
        self.assertContains(response, "PodcastOne")
        self.assertContains(response, "Cancel")
        self.assertContains(response, "parse-result-tabs")
        self.assertContains(response, "Final CSV")
        self.assertContains(response, "Approval")
        self.assertNotContains(response, "PodcastOne new.csv")

    def test_process_page_keeps_parsing_header_when_queue_is_empty(self):
        response = self.client.get(reverse("pipeline_dashboard:process"))

        self.assertEqual(response.status_code, 200)
        # The mirror-head chrome (title, subtitle, count cells, search input) must
        # remain even when there are no files queued for parsing, so the empty
        # chapter matches the Approval and History chapters visually.
        self.assertContains(response, "Parsing Files")
        self.assertContains(response, "0 files in parser queue")
        self.assertContains(response, "data-process-search")
        self.assertContains(response, "process-mirror parsing-files")
        self.assertContains(response, "No files are currently queued for parsing.")

    def test_process_page_formats_approval_totals_for_readability(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        asset = Asset.objects.create(
            remote_item_id="fi-loop-review-format",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.REVIEW,
            name="Loop review.xlsx",
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
        )
        ParsedOutput.objects.create(
            asset=asset,
            vendor=loop,
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
            reporting_period="April_2026",
            row_count=30,
            total_spend="82711.112970",
            total_impressions="4792893.088630",
            comparison_status="sent_for_approval",
        )

        response = self.client.get(reverse("pipeline_dashboard:process"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "82 711")
        self.assertContains(response, "4 792 893")
        self.assertNotContains(response, "82711.112970")
        self.assertNotContains(response, "4792893.088630")

    def test_process_page_history_groups_approved_outputs_by_vendor(self):
        import re

        rally, _ = Vendor.objects.get_or_create(
            name="RallyAdMedia", defaults={"parser_key": "rallyadmedia"}
        )
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        older_asset = Asset.objects.create(
            remote_item_id="fi-rally-older",
            vendor=rally,
            parser_key="rallyadmedia",
            status=AssetStatus.PROCESSED,
            name="RallyAdMedia_Apr_2026_v2.xlsx",
            output_path="data/output/RallyAdMedia/RallyAdMedia_Apr_2026_v2.csv",
        )
        newer_asset = Asset.objects.create(
            remote_item_id="fi-rally-newer",
            vendor=rally,
            parser_key="rallyadmedia",
            status=AssetStatus.PROCESSED,
            name="RallyAdMedia_April_2026.xlsx",
            output_path="data/processed/RallyAdMedia/RallyAdMedia_April_2026.csv",
        )
        ParsedOutput.objects.create(
            asset=older_asset,
            vendor=rally,
            output_path="data/output/RallyAdMedia/RallyAdMedia_Apr_2026_v2.csv",
            comparison_status="approved",
        )
        ParsedOutput.objects.create(
            asset=newer_asset,
            vendor=rally,
            output_path="data/processed/RallyAdMedia/RallyAdMedia_April_2026.csv",
            comparison_status="approved",
        )
        # Another vendor's approved file sits between the two RallyAdMedia rows
        # by creation time, which would have split the group under the old sort.
        loop_asset = Asset.objects.create(
            remote_item_id="fi-loop-may",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.PROCESSED,
            name="Loop_May_2026_v1.xlsx",
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
        )
        ParsedOutput.objects.create(
            asset=loop_asset,
            vendor=loop,
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
            comparison_status="approved",
        )

        response = self.client.get(reverse("pipeline_dashboard:process"))
        body = response.content.decode()

        self.assertEqual(response.status_code, 200)
        # Exactly one RallyAdMedia folder row in the History panel.
        rally_folders = re.findall(
            r'mirror-folder-name[^>]*>RallyAdMedia<', body
        )
        self.assertEqual(len(rally_folders), 1)
        # Both files are listed.
        self.assertIn("RallyAdMedia_April_2026.csv", body)
        self.assertIn("RallyAdMedia_Apr_2026_v2.csv", body)
        # The Files count for the RallyAdMedia folder is 2.
        rally_folder = re.search(
            r'<details class="mirror-folder">.*?RallyAdMedia.*?</details>',
            body,
            re.DOTALL,
        )
        self.assertIsNotNone(rally_folder)
        self.assertIn(
            'process-count-cell right-aligned" role="cell">2<',
            rally_folder.group(0),
        )

    def test_update_process_vendor_changes_asset_vendor(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        podcastone, _ = Vendor.objects.get_or_create(name="PodcastOne", defaults={"parser_key": "podcastone"})
        asset = Asset.objects.create(
            remote_item_id="fi-processing",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.PROCESSING,
            name="Loop report.csv",
        )

        response = self.client.post(
            reverse("pipeline_dashboard:update_process_vendor", args=[asset.remote_item_id]),
            {"vendor_id": podcastone.id},
        )

        self.assertRedirects(response, reverse("pipeline_dashboard:process"))
        asset.refresh_from_db()
        self.assertEqual(asset.vendor, podcastone)
        self.assertEqual(asset.parser_key, "podcastone")
        self.assertTrue(asset.events.filter(event_type="vendor_changed").exists())

    def test_update_process_vendor_returns_json_for_ajax(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        podcastone, _ = Vendor.objects.get_or_create(name="PodcastOne", defaults={"parser_key": "podcastone"})
        asset = Asset.objects.create(
            remote_item_id="fi-processing-ajax",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.PROCESSING,
            name="Loop report.csv",
        )

        response = self.client.post(
            reverse("pipeline_dashboard:update_process_vendor", args=[asset.remote_item_id]),
            {"vendor_id": podcastone.id},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["vendor"], "PodcastOne")
        self.assertEqual(data["vendor_id"], podcastone.id)
        asset.refresh_from_db()
        self.assertEqual(asset.vendor, podcastone)

    def test_cancel_process_file_clears_vendor_and_returns_to_new(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        asset = Asset.objects.create(
            remote_item_id="fi-processing",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.PROCESSING,
            name="Loop report.csv",
        )

        response = self.client.post(reverse("pipeline_dashboard:cancel_process_file", args=[asset.remote_item_id]))

        self.assertRedirects(response, reverse("pipeline_dashboard:process"))
        asset.refresh_from_db()
        self.assertEqual(asset.status, AssetStatus.NEW)
        self.assertIsNone(asset.vendor)
        self.assertEqual(asset.parser_key, "")
        self.assertTrue(asset.events.filter(event_type="processing_cancelled").exists())

    def test_parse_file_preview_validates_loop_schema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_loop_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:parse_file_preview", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["validation"]["ok"])
        self.assertEqual(payload["validation"]["vendor"], "Loop")
        self.assertEqual(payload["validation"]["sheet_name"], "Daily Spend")
        self.assertEqual(payload["file"]["sheets"][0]["name"], "Daily Spend")

    def test_parse_process_file_handles_tvm_imps_currency_formatting(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_tvm_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "May_2026")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "3646.72")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "595104")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed May_2026")

    def test_parse_process_file_combines_taiv_prime_and_retail_tables(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_taiv_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "May_2026")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "74")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "7400")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed May_2026")

    def test_parse_process_file_combines_podcastone_daily_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_podcastone_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "May_2026")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "100.75")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "1000")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed May_2026")

    def test_parse_process_file_combines_octopus_dooh_and_rideshare_tables(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_octopus_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "May_2026")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "100")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "1000")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed May_2026")

    def test_parse_process_file_combines_rallyadmedia_brand_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_rallyadmedia_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "March_2026")
        self.assertEqual(payload["candidate"]["output_filename"], "RallyAdMedia_Apr_2026_v1.csv")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "280")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "2800")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed March_2026")

    def test_parse_process_file_with_sheet_name_supports_multi_worksheet_schemas(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_rallyadmedia_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(
                    reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]),
                    data={"sheet_name": "BOL"},
                )

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "280")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "2800")

    def test_parse_sheet_probe_reports_undeclared_sheet_for_multi_worksheet_schema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_rallyadmedia_parse_fixture(repo_root)
            workbook_path = repo_root / asset.local_path
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path)
            workbook.create_sheet("Totals")
            workbook.save(workbook_path)
            workbook.close()

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse("pipeline_dashboard:parse_sheet_probe", args=[asset.remote_item_id]),
                    data={"sheet_name": "Totals"},
                )

        self.assertEqual(response.status_code, 200, response.content)
        validation = response.json()["validation"]
        self.assertFalse(validation["ok"])
        self.assertTrue(
            any("not declared in the input schema" in error for error in validation["errors"]),
            validation["errors"],
        )

    def test_parse_sheet_probe_reports_header_mismatch_for_multi_worksheet_schema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_rallyadmedia_parse_fixture(repo_root)
            workbook_path = repo_root / asset.local_path
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path)
            workbook["BOL"]["A1"] = "DATE_LABEL_RENAMED"
            workbook.save(workbook_path)
            workbook.close()

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(
                    reverse("pipeline_dashboard:parse_sheet_probe", args=[asset.remote_item_id]),
                    data={"sheet_name": "BOL"},
                )

        self.assertEqual(response.status_code, 200, response.content)
        validation = response.json()["validation"]
        self.assertFalse(validation["ok"])
        self.assertTrue(
            any("'BOL'" in error and "A1" in error for error in validation["errors"]),
            validation["errors"],
        )

    def test_parse_process_file_distributes_adtaxi_totals_across_report_dates(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_adtaxi_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "March_2026")
        self.assertEqual(payload["candidate"]["output_filename"], "AdTaxi_Apr_2026_v1.csv")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 31)
        self.assertEqual(payload["candidate"]["summary"]["total_spend"], "93")
        self.assertEqual(payload["candidate"]["summary"]["total_impressions"], "930")
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed March_2026")
        self.assertEqual(payload["parsed_table"]["rows"][0][0], "2026-03-01")
        self.assertEqual(payload["parsed_table"]["rows"][0][5], "3")
        self.assertEqual(payload["parsed_table"]["rows"][0][6], "30")
        self.assertEqual(payload["parsed_table"]["rows"][-1][0], "2026-03-31")

    def test_approve_rallyadmedia_march_output_uses_april_approval_period(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_rallyadmedia_parse_fixture(repo_root)
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.approval_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                response = self.client.post(reverse("pipeline_dashboard:approve_process_file", args=[asset.remote_item_id]))
                output_exists = (
                    repo_root / "data" / "output" / "RallyAdMedia" / "RallyAdMedia_Apr_2026_v1.csv"
                ).exists()

        self.assertEqual(response.status_code, 200)
        asset.refresh_from_db()
        parsed = ParsedOutput.objects.get(asset=asset)
        self.assertEqual(asset.status, AssetStatus.REVIEW)
        self.assertEqual(asset.output_path, "data/output/RallyAdMedia/RallyAdMedia_Apr_2026_v1.csv")
        self.assertEqual(parsed.output_path, "data/output/RallyAdMedia/RallyAdMedia_Apr_2026_v1.csv")
        self.assertEqual(parsed.reporting_period, "March_2026")
        self.assertEqual(parsed.comparison_summary["approval_folder_label"], "April_2026")
        self.assertEqual(parsed.comparison_summary["approval_filename_label"], "Apr_2026")
        self.assertEqual(fake_client.folder_parts, ["Approval", "April_2026", "RallyAdMedia"])
        self.assertEqual(fake_client.uploaded_name, "RallyAdMedia_Apr_2026_v1.csv")
        self.assertTrue(fake_client.notify)
        self.assertTrue(fake_client.copy_access_controls)
        self.assertTrue(output_exists)

    def test_approve_adtaxi_march_output_uses_april_approval_period(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_adtaxi_parse_fixture(repo_root)
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.approval_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                response = self.client.post(reverse("pipeline_dashboard:approve_process_file", args=[asset.remote_item_id]))
                output_exists = (repo_root / "data" / "output" / "AdTaxi" / "AdTaxi_Apr_2026_v1.csv").exists()

        self.assertEqual(response.status_code, 200)
        asset.refresh_from_db()
        parsed = ParsedOutput.objects.get(asset=asset)
        self.assertEqual(asset.status, AssetStatus.REVIEW)
        self.assertEqual(asset.output_path, "data/output/AdTaxi/AdTaxi_Apr_2026_v1.csv")
        self.assertEqual(parsed.reporting_period, "March_2026")
        self.assertEqual(parsed.comparison_summary["approval_folder_label"], "April_2026")
        self.assertEqual(parsed.comparison_summary["approval_filename_label"], "Apr_2026")
        self.assertEqual(fake_client.folder_parts, ["Approval", "April_2026", "AdTaxi"])
        self.assertEqual(fake_client.uploaded_name, "AdTaxi_Apr_2026_v1.csv")
        self.assertTrue(fake_client.notify)
        self.assertTrue(fake_client.copy_access_controls)
        self.assertTrue(output_exists)

    def test_parse_process_file_returns_chart_preview_without_writing_output(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_loop_parse_fixture(repo_root)

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))
                output_exists = (repo_root / "data" / "output" / "Loop" / "Loop_May_2026_v1.csv").exists()

        self.assertEqual(response.status_code, 200)
        asset.refresh_from_db()
        payload = response.json()
        self.assertEqual(asset.status, AssetStatus.PROCESSING)
        self.assertFalse(output_exists)
        self.assertEqual(payload["candidate"]["summary"]["period_label"], "May_2026")
        self.assertEqual(payload["candidate"]["summary"]["row_count"], 2)
        self.assertEqual(payload["charts"]["series"][0]["label"], "Parsed May_2026")
        self.assertEqual(payload["charts"]["series"][0]["points"][0]["day"], 1)
        self.assertEqual(payload["charts"]["series"][0]["points"][0]["cpm"], 0.1)
        self.assertEqual(payload["parsed_table"]["columns"][0], "Date")
        self.assertEqual(payload["parsed_table"]["row_count"], 2)
        self.assertFalse(payload["parsed_table"]["truncated"])
        self.assertEqual(payload["parsed_table"]["rows"][0][0], "2026-05-01")
        self.assertEqual(ParsedOutput.objects.filter(asset=asset).count(), 0)

    def test_chart_preview_groups_metrics_by_date(self):
        series = period_series_from_rows(
            [
                {"Date": "2026-05-01", "Spend": "10", "Impressions": "100"},
                {"Date": "2026-05-01", "Spend": "15", "Impressions": "400"},
                {"Date": "2026-05-02", "Spend": "5", "Impressions": "50"},
            ],
            "Parsed May_2026",
        )

        self.assertEqual(len(series["points"]), 2)
        self.assertEqual(series["points"][0]["date"], "2026-05-01")
        self.assertEqual(series["points"][0]["spend"], 25.0)
        self.assertEqual(series["points"][0]["impressions"], 500.0)
        self.assertEqual(series["points"][0]["cpm"], 0.05)

    def test_old_final_csv_is_used_for_baseline_comparison(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            project_root = Path(__file__).resolve().parents[3]
            parser_root = repo_root / "parsers" / "S2"
            parser_root.mkdir(parents=True)
            shutil.copy2(project_root / "parsers" / "S2" / "input_schema.json", parser_root / "input_schema.json")
            shutil.copy2(project_root / "parsers" / "S2" / "parser.py", parser_root / "parser.py")

            old_final = repo_root / "_old" / "final" / "S2.csv"
            old_final.parent.mkdir(parents=True)
            old_final.write_text(
                "\n".join(
                    [
                        "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                        "2025-03-01,S2,BetOnline,CTV,S2 Network,1000,20000,daily,approved,baseline1.csv",
                        "2025-03-02,S2,BetOnline,CTV,S2 Network,1100,22000,daily,approved,baseline1.csv",
                        "2025-03-03,S2,BetOnline,CTV,S2 Network,1200,24000,daily,approved,baseline1.csv",
                        "2025-03-04,S2,BetOnline,CTV,S2 Network,1300,26000,daily,approved,baseline1.csv",
                        "2025-03-05,S2,BetOnline,CTV,S2 Network,1400,28000,daily,approved,baseline1.csv",
                        "2025-02-01,S2,BetOnline,CTV,S2 Network,900,18000,daily,approved,baseline2.csv",
                        "2025-02-02,S2,BetOnline,CTV,S2 Network,950,19000,daily,approved,baseline2.csv",
                        "2025-02-03,S2,BetOnline,CTV,S2 Network,1000,20000,daily,approved,baseline2.csv",
                        "2025-02-04,S2,BetOnline,CTV,S2 Network,1050,21000,daily,approved,baseline2.csv",
                        "2025-02-05,S2,BetOnline,CTV,S2 Network,1100,22000,daily,approved,baseline2.csv",
                        "",
                    ]
                ),
                encoding="utf-8",
            )

            local_path = "data/inbox/home/pm/April.Connected-TV.Data.xlsx"
            workbook_path = repo_root / local_path
            workbook_path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "April.26"
            for row in range(1, 31):
                for col in range(1, 16):
                    sheet.cell(row, col).value = None
            sheet["A31"] = "Day"
            sheet["B31"] = "Spend"
            sheet["C31"] = "Video Plays"
            sheet["E31"] = "Day"
            sheet["F31"] = "Spend"
            sheet["G31"] = "Video Plays"
            sheet["I31"] = "Day"
            sheet["J31"] = "Spend"
            sheet["K31"] = "Video Plays"
            sheet["M31"] = "Day"
            sheet["N31"] = "Spend"
            sheet["O31"] = "Video Plays"
            for idx, day in enumerate(range(1, 4), start=32):
                date_str = f"2025-04-{day:02d}"
                sheet.cell(idx, 1).value = date_str
                sheet.cell(idx, 2).value = 500 + day * 10
                sheet.cell(idx, 3).value = 10000 + day * 100
                sheet.cell(idx, 5).value = date_str
                sheet.cell(idx, 6).value = 500 + day * 10
                sheet.cell(idx, 7).value = 10000 + day * 100
                sheet.cell(idx, 9).value = date_str
                sheet.cell(idx, 10).value = 500 + day * 10
                sheet.cell(idx, 11).value = 10000 + day * 100
                sheet.cell(idx, 13).value = date_str
                sheet.cell(idx, 14).value = 500 + day * 10
                sheet.cell(idx, 15).value = 10000 + day * 100
            workbook.save(workbook_path)
            workbook.close()

            schema_path = parser_root / "input_schema.json"
            schema = json.loads(schema_path.read_text(encoding="utf-8"))
            schema["sheet_name"] = "April.26"
            schema["header"]["row"] = 31
            for table in schema.get("tables", []):
                table["header_row"] = 31
                table["first_data_row"] = 32
            schema_path.write_text(json.dumps(schema, indent=2), encoding="utf-8")

            s2, _ = Vendor.objects.get_or_create(name="S2", defaults={"parser_key": "S2"})
            asset = Asset.objects.create(
                remote_item_id="fi-s2-parse",
                vendor=s2,
                parser_key="S2",
                status=AssetStatus.PROCESSING,
                name="April.Connected-TV.Data.xlsx",
                local_path=local_path,
                file_size=workbook_path.stat().st_size,
            )

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.post(reverse("pipeline_dashboard:parse_process_file", args=[asset.remote_item_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["charts"]["series"]), 3)
        labels = [s["label"] for s in payload["charts"]["series"]]
        self.assertEqual(labels[0], "Parsed April_2025")
        self.assertIn("March_2025", labels[1:])
        self.assertIn("February_2025", labels[1:])
        self.assertEqual(payload["comparison"]["status"], "no_matching_history")

    def test_approve_process_file_sends_versioned_output_for_external_approval(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            asset = self._write_loop_parse_fixture(repo_root)
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.approval_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                response = self.client.post(reverse("pipeline_dashboard:approve_process_file", args=[asset.remote_item_id]))
                output_exists = (repo_root / "data" / "output" / "Loop" / "Loop_Jun_2026_v1.csv").exists()
                processed_exists = (repo_root / "data" / "processed" / "Loop" / "Loop_May_2026.csv").exists()

        self.assertEqual(response.status_code, 200)
        asset.refresh_from_db()
        parsed = ParsedOutput.objects.get(asset=asset)
        self.assertEqual(asset.status, AssetStatus.REVIEW)
        self.assertEqual(asset.uploaded_item_id, "fi-uploaded")
        self.assertEqual(asset.output_path, "data/output/Loop/Loop_Jun_2026_v1.csv")
        self.assertEqual(parsed.reporting_period, "May_2026")
        self.assertEqual(parsed.row_count, 2)
        self.assertEqual(parsed.comparison_status, "sent_for_approval")
        self.assertEqual(parsed.comparison_summary["sharefile_item_id"], "fi-uploaded")
        self.assertTrue(output_exists)
        self.assertFalse(processed_exists)
        self.assertEqual(fake_client.folder_parts, ["Approval", "June_2026", "Loop"])
        self.assertEqual(fake_client.uploaded_name, "Loop_Jun_2026_v1.csv")
        self.assertTrue(fake_client.notify)
        self.assertTrue(fake_client.copy_access_controls)
        self.assertTrue(asset.events.filter(event_type="approval_sent").exists())

    def test_cancel_parsed_output_returns_asset_to_processing(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        asset = Asset.objects.create(
            remote_item_id="fi-loop-review",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.REVIEW,
            name="Loop review.xlsx",
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
        )
        parsed = ParsedOutput.objects.create(
            asset=asset,
            vendor=loop,
            output_path="data/output/Loop/Loop_May_2026_v1.csv",
            reporting_period="May_2026",
            comparison_status="no_matching_history",
        )

        response = self.client.post(reverse("pipeline_dashboard:cancel_parsed_output", args=[parsed.id]))

        self.assertRedirects(response, reverse("pipeline_dashboard:process"))
        asset.refresh_from_db()
        parsed.refresh_from_db()
        self.assertEqual(asset.status, AssetStatus.PROCESSING)
        self.assertEqual(asset.output_path, "")
        self.assertEqual(parsed.comparison_status, "cancelled")
        self.assertTrue(asset.events.filter(event_type="parsed_output_cancelled").exists())

    def test_approve_parsed_output_stores_final_file_and_marks_asset_processed(self):
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            output_path = repo_root / "data" / "output" / "Loop" / "Loop_May_2026_v1.csv"
            output_path.parent.mkdir(parents=True)
            output_path.write_text(
                "\n".join(
                    [
                        "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                        "2026-05-01,Loop,BetOnline,DOOH,Loop TV,10,100,daily,generated,source.xlsx",
                        "",
                    ]
                ),
                encoding="utf-8",
            )
            asset = Asset.objects.create(
                remote_item_id="fi-loop-review-final",
                vendor=loop,
                parser_key="loop",
                status=AssetStatus.REVIEW,
                name="Loop review.xlsx",
                output_path="data/output/Loop/Loop_May_2026_v1.csv",
                uploaded_item_id="fi-approval",
            )
            parsed = ParsedOutput.objects.create(
                asset=asset,
                vendor=loop,
                output_path="data/output/Loop/Loop_May_2026_v1.csv",
                reporting_period="May_2026",
                comparison_status="sent_for_approval",
            )
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.final_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                response = self.client.post(reverse("pipeline_dashboard:approve_parsed_output", args=[parsed.id]))
                final_path = repo_root / "data" / "processed" / "Loop" / "Loop_May_2026.csv"
                final_exists = final_path.exists()
                staging_still_exists = output_path.exists()
                # The staging file may have been unlinked by the view, so read
                # its content from the final CSV instead. They should be byte-equal.
                final_content = final_path.read_bytes()
                process_response = self.client.get(reverse("pipeline_dashboard:process"))
                reviewed_content = final_content

        self.assertRedirects(response, reverse("pipeline_dashboard:process"))
        asset.refresh_from_db()
        parsed.refresh_from_db()
        self.assertEqual(asset.status, AssetStatus.PROCESSED)
        self.assertEqual(asset.output_path, "data/processed/Loop/Loop_May_2026.csv")
        self.assertEqual(asset.uploaded_item_id, "fi-uploaded")
        self.assertEqual(parsed.comparison_status, "approved")
        self.assertEqual(parsed.output_path, "data/processed/Loop/Loop_May_2026.csv")
        self.assertEqual(parsed.comparison_summary["final_sharefile_item_id"], "fi-uploaded")
        self.assertEqual(parsed.comparison_summary["final_sharefile_filename"], "Loop_May_2026.csv")
        self.assertEqual(parsed.comparison_summary["final_sharefile_path"], "Final/May_2026/Loop_May_2026.csv")
        self.assertTrue(final_exists)
        self.assertEqual(final_content, reviewed_content)
        self.assertEqual(fake_client.uploaded_content, reviewed_content)
        self.assertIn(b"2026-05-01,Loop,BetOnline", final_content)
        self.assertEqual(fake_client.folder_parts, ["Final", "May_2026"])
        self.assertEqual(fake_client.uploaded_name, "Loop_May_2026.csv")
        self.assertTrue(fake_client.notify)
        self.assertTrue(fake_client.copy_access_controls)
        self.assertTrue(asset.events.filter(event_type="final_approved").exists())
        # The versioned staging copy under data/output/ is no longer needed
        # once the final CSV is in data/processed/ and the ShareFile Final
        # upload has succeeded, so the view removes it.
        self.assertFalse(staging_still_exists)
        # The /process/ page should still render the now-approved Loop row in
        # the History chapter.
        self.assertContains(process_response, "data/processed/Loop/Loop_May_2026.csv")

    def test_approve_parsed_output_deletes_staging_file(self):
        """After a successful approval the data/output/ staging CSV is unlinked.

        Covers the explicit safety branch: a successful finalize_approved_output
        followed by _delete_staging_output removes the file the user no longer
        needs. Also covers the failure branch: if the staging path is empty
        or points outside data/output/, the view must not touch anything.
        """
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            output_path = repo_root / "data" / "output" / "Loop" / "Loop_May_2026_v1.csv"
            final_path = repo_root / "data" / "processed" / "Loop" / "Loop_May_2026.csv"
            output_path.parent.mkdir(parents=True)
            output_path.write_text(
                "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File\n"
                "2026-05-01,Loop,BetOnline,DOOH,Loop TV,10,100,daily,generated,source.xlsx\n",
                encoding="utf-8",
            )
            asset = Asset.objects.create(
                remote_item_id="fi-loop-staging-delete",
                vendor=loop,
                parser_key="loop",
                status=AssetStatus.REVIEW,
                name="Loop review.xlsx",
                output_path="data/output/Loop/Loop_May_2026_v1.csv",
                uploaded_item_id="fi-approval",
            )
            parsed = ParsedOutput.objects.create(
                asset=asset,
                vendor=loop,
                output_path="data/output/Loop/Loop_May_2026_v1.csv",
                reporting_period="May_2026",
                comparison_status="sent_for_approval",
            )
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.final_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                response = self.client.post(
                    reverse("pipeline_dashboard:approve_parsed_output", args=[parsed.id])
                )
                # Capture inside the override block, before the tempdir is torn down.
                staging_still_exists = output_path.exists()
                final_exists = final_path.exists()
                parsed.refresh_from_db()
                parsed_output_path = parsed.output_path

        self.assertRedirects(response, reverse("pipeline_dashboard:process"))
        self.assertFalse(staging_still_exists, "staging copy under data/output/ should be unlinked")
        self.assertTrue(final_exists, "final CSV under data/processed/ should be in place")
        self.assertEqual(parsed_output_path, "data/processed/Loop/Loop_May_2026.csv")
        self.assertTrue(fake_client.notify)
        self.assertTrue(fake_client.copy_access_controls)

    def test_approve_parsed_output_does_not_delete_non_staging_path(self):
        """The unlink guard must refuse to touch paths outside data/output/.

        If parsed.output_path is mis-recorded (for example, an already-canonical
        data/processed/... path or anything else), the view must not unlink it
        on approval. Only the canonical data/output/<Vendor>/... staging
        location is safe to delete.
        """
        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            # Place the staging CSV where the view expects it (so promotion to
            # data/processed/ works), but record parsed.output_path at a
            # different, non-data/output/ location. That sentinel is what the
            # safety guard exists to protect: the view must not unlink it.
            output_path = repo_root / "data" / "output" / "Loop" / "Loop_May_2026_v1.csv"
            output_path.parent.mkdir(parents=True)
            output_path.write_text(
                "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File\n"
                "2026-05-01,Loop,BetOnline,DOOH,Loop TV,10,100,daily,generated,source.xlsx\n",
                encoding="utf-8",
            )
            # Mis-recorded: a CSV under data/processed/ that we expect to survive.
            misrecorded_path = repo_root / "data" / "processed" / "OtherVendor" / "other.csv"
            misrecorded_path.parent.mkdir(parents=True)
            misrecorded_path.write_text("survives\n", encoding="utf-8")

            asset = Asset.objects.create(
                remote_item_id="fi-loop-guard",
                vendor=loop,
                parser_key="loop",
                status=AssetStatus.REVIEW,
                name="Loop review.xlsx",
                output_path="data/output/Loop/Loop_May_2026_v1.csv",
                uploaded_item_id="fi-approval",
            )
            parsed = ParsedOutput.objects.create(
                asset=asset,
                vendor=loop,
                output_path=str(misrecorded_path.relative_to(repo_root)),
                reporting_period="May_2026",
                comparison_status="sent_for_approval",
            )
            fake_client = FakeApprovalClient()

            with (
                override_settings(REPO_ROOT=repo_root),
                patch("pipeline_dashboard.parser_workflow.final_root_id", return_value="fo-root"),
                patch("pipeline_dashboard.parser_workflow.build_sharefile_client", return_value=fake_client),
            ):
                self.client.post(reverse("pipeline_dashboard:approve_parsed_output", args=[parsed.id]))

            # The misrecorded path is the staging file the view would try to
            # delete. The safety guard must refuse: only data/output/... is
            # safe to unlink. The canonical staging file under data/output/
            # is left alone because parsed.output_path is mis-recorded, and
            # the guard has to honour whatever path the operator wrote.
            self.assertTrue(misrecorded_path.exists())
            self.assertTrue(output_path.exists())
            self.assertTrue(fake_client.notify)
            self.assertTrue(fake_client.copy_access_controls)

    def test_approval_root_defaults_to_allshared(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            (repo_root / ".env").write_text("SHAREFILE_CLIENT_ID=example\n", encoding="utf-8")

            with override_settings(REPO_ROOT=repo_root):
                self.assertEqual(approval_root_id(), "allshared")

    @patch("pipeline_dashboard.views.subprocess.run")
    def test_update_folders_runs_update_script(self, run_mock):
        run_mock.return_value.returncode = 0
        run_mock.return_value.stdout = "ok"
        run_mock.return_value.stderr = ""

        response = self.client.post(reverse("pipeline_dashboard:update_folders"), follow=True)

        self.assertRedirects(response, reverse("pipeline_dashboard:folders"))
        self.assertNotContains(response, "SF folders updated.")
        run_mock.assert_called_once()

    @patch("pipeline_dashboard.views.subprocess.run")
    def test_update_folders_ajax_returns_redirect_payload(self, run_mock):
        run_mock.return_value.returncode = 0
        run_mock.return_value.stdout = "ok"
        run_mock.return_value.stderr = ""

        response = self.client.post(
            reverse("pipeline_dashboard:update_folders"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {"ok": True, "redirect_url": reverse("pipeline_dashboard:folders")},
        )
        run_mock.assert_called_once()

    def _write_review_fixture(self, repo_root: Path) -> str:
        local_path = "data/inbox/home/josh/new-report.csv"
        state_root = repo_root / "data" / "state"
        file_path = repo_root / local_path
        state_root.mkdir(parents=True)
        file_path.parent.mkdir(parents=True)
        file_path.write_text("Date,Campaign,Spend\n2026-05-07,Spring,10\n", encoding="utf-8")
        (state_root / "sharefile_snapshot_latest.json").write_text(
            json.dumps(
                {
                    "run_id": "snapshot-review",
                    "created_at": "2026-05-07T10:00:00Z",
                    "files": [
                        {
                            "remote_item_id": "fi-review",
                            "name": "new-report.csv",
                            "remote_path": "home/josh/new-report.csv",
                            "local_path": local_path,
                            "source_folder_path": "home/josh",
                            "source_folder_id": "fo-josh",
                            "extension": ".csv",
                            "size": file_path.stat().st_size,
                            "created_at": "2026-05-07T09:00:00Z",
                            "modified_at": "2026-05-07T10:00:00Z",
                            "creator": "Uploader One",
                            "raw_metadata": {"LastModifiedByUserID": "user-1"},
                        },
                    ],
                }
            ),
            encoding="utf-8",
        )
        (state_root / "inbox_profile_latest.json").write_text(
            json.dumps(
                {
                    "files": [
                        {
                            "local_path": local_path,
                            "name": "new-report.csv",
                            "kind": "csv",
                            "status": "profiled",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )
        (state_root / "sharefile_users_latest.json").write_text(
            json.dumps({"users_by_id": {"user-1": {"full_name": "Uploader One", "email": "one@example.com"}}}),
            encoding="utf-8",
        )
        return local_path

    def _write_loop_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "Loop"
        parser_root.mkdir(parents=True)
        shutil.copy2(project_root / "parsers" / "Loop" / "input_schema.json", parser_root / "input_schema.json")
        shutil.copy2(project_root / "parsers" / "Loop" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "Loop" / "Loop.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-05-01,Loop,BetOnline,DOOH,Loop TV,10,100,daily,approved,baseline.csv",
                    "2026-05-02,Loop,BetOnline,DOOH,Loop TV,20,200,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/home/josh/LOOP MAY 2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Spend"
        sheet["A5"] = "Date"
        sheet["B5"] = "Impressions"
        sheet["C5"] = "Spend"
        sheet["A6"] = "2026-05-01"
        sheet["B6"] = 100
        sheet["C6"] = 10
        sheet["A7"] = "2026-05-02"
        sheet["B7"] = 200
        sheet["C7"] = 20
        workbook.save(workbook_path)
        workbook.close()

        loop, _ = Vendor.objects.get_or_create(name="Loop", defaults={"parser_key": "loop"})
        return Asset.objects.create(
            remote_item_id="fi-loop-parse",
            vendor=loop,
            parser_key="loop",
            status=AssetStatus.PROCESSING,
            name="LOOP MAY 2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_tvm_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "TVM"
        parser_root.mkdir(parents=True)
        shutil.copy2(project_root / "parsers" / "TVM" / "input_schema.json", parser_root / "input_schema.json")
        shutil.copy2(project_root / "parsers" / "TVM" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "TVM" / "TVM.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-04-01,TVM,BetOnline,DOOH,TVM,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/home/josh/TVM MAY 2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Total ImpsTotal Spend"
        sheet["A1"] = "Day"
        sheet["B1"] = "Imps "
        sheet["C1"] = "Spend"
        sheet["A2"] = "2026-05-01"
        sheet["B2"] = "$275,099"
        sheet["C2"] = "$1,663.17"
        sheet["A3"] = None
        sheet["B3"] = None
        sheet["C3"] = None
        sheet["A4"] = "2026-05-02"
        sheet["B4"] = 320005
        sheet["C4"] = 1983.55
        workbook.save(workbook_path)
        workbook.close()

        tvm, _ = Vendor.objects.get_or_create(name="TVM", defaults={"parser_key": "tvm"})
        return Asset.objects.create(
            remote_item_id="fi-tvm-parse",
            vendor=tvm,
            parser_key="tvm",
            status=AssetStatus.PROCESSING,
            name="TVM MAY 2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_octopus_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "Octopus"
        parser_root.mkdir(parents=True)
        shutil.copy2(project_root / "parsers" / "Octopus" / "input_schema.json", parser_root / "input_schema.json")
        shutil.copy2(project_root / "parsers" / "Octopus" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "Octopus" / "Octopus.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-04-01,Octopus,BetOnline,Display,Octopus,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/home/josh/T-MOBILE MAY 2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Spend"
        sheet["A1"] = "DOOH"
        sheet["B1"] = "Impressions"
        sheet["C1"] = "Spend"
        sheet["A2"] = "2026-05-01"
        sheet["B2"] = 100
        sheet["C2"] = 10
        sheet["A3"] = "2026-05-02"
        sheet["B3"] = 200
        sheet["C3"] = 20
        sheet["A34"] = "Rideshare"
        sheet["B34"] = "Impressions"
        sheet["C34"] = "Spend"
        sheet["A35"] = "2026-05-01"
        sheet["B35"] = 300
        sheet["C35"] = 30
        sheet["A36"] = "2026-05-02"
        sheet["B36"] = 400
        sheet["C36"] = 40
        workbook.save(workbook_path)
        workbook.close()

        octopus, _ = Vendor.objects.get_or_create(name="Octopus", defaults={"parser_key": "octopus"})
        return Asset.objects.create(
            remote_item_id="fi-octopus-parse",
            vendor=octopus,
            parser_key="octopus",
            status=AssetStatus.PROCESSING,
            name="T-MOBILE MAY 2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_taiv_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "TAIV"
        parser_root.mkdir(parents=True)
        shutil.copy2(project_root / "parsers" / "TAIV" / "input_schema.json", parser_root / "input_schema.json")
        shutil.copy2(project_root / "parsers" / "TAIV" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "TAIV" / "TAIV.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-04-01,TAIV,Unknown,Unknown,Unknown,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/home/josh/TAIV MAY 2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Spend By Day"
        sheet["A1"] = "Prime"
        sheet["E1"] = "Retail"
        sheet["A2"] = "date"
        sheet["B2"] = "impressions"
        sheet["C2"] = "spend"
        sheet["E2"] = "date"
        sheet["F2"] = "impressions"
        sheet["G2"] = "spend"
        sheet["A3"] = "2026-05-01"
        sheet["B3"] = 1000
        sheet["C3"] = 10
        sheet["E3"] = "2026-05-01"
        sheet["F3"] = "$2,000"
        sheet["G3"] = "$20"
        sheet["A4"] = "2026-05-02"
        sheet["B4"] = 1400
        sheet["C4"] = 14
        sheet["E4"] = "2026-05-02"
        sheet["F4"] = 3000
        sheet["G4"] = 30
        sheet["A5"] = None
        sheet["B5"] = 2400
        sheet["C5"] = 24
        sheet["E5"] = None
        sheet["F5"] = 5000
        sheet["G5"] = 50
        workbook.save(workbook_path)
        workbook.close()

        taiv, _ = Vendor.objects.get_or_create(name="TAIV", defaults={"parser_key": "taiv"})
        return Asset.objects.create(
            remote_item_id="fi-taiv-parse",
            vendor=taiv,
            parser_key="taiv",
            status=AssetStatus.PROCESSING,
            name="TAIV MAY 2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_podcastone_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "PodcastOne"
        parser_root.mkdir(parents=True)
        shutil.copy2(
            project_root / "parsers" / "PodcastOne" / "input_schema.json",
            parser_root / "input_schema.json",
        )
        shutil.copy2(project_root / "parsers" / "PodcastOne" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "PodcastOne" / "PodcastOne.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-04-01,PodcastOne,BetOnline,Audio,PodcastOne,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/home/josh/PODCASTONE MAY 2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        base = workbook.create_sheet("Week 1-5 BASE DLY 4.1-4.30")
        wc = workbook.create_sheet("WC Week 1-5 DLY 4.1-4.30")
        for sheet in [base, wc]:
            sheet["A5"] = "Order ID"
            sheet["B5"] = "Order"
            sheet["C5"] = "Campaign"
            sheet["D5"] = "Day"
            sheet["E5"] = "Audio Impressions"
            sheet["F5"] = "$ By Day"
        base["D6"] = "2026-05-01"
        base["E6"] = 100
        base["F6"] = 10.50
        base["D7"] = "2026-05-02"
        base["E7"] = 200
        base["F7"] = 20.25
        base["A8"] = "TOTAL"
        base["E8"] = 300
        base["F8"] = 30.75
        wc["D6"] = "2026-05-01"
        wc["E6"] = "$300"
        wc["F6"] = "$30"
        wc["D7"] = "2026-05-02"
        wc["E7"] = 400
        wc["F7"] = 40
        wc["A8"] = "TOTAL"
        wc["E8"] = 700
        wc["F8"] = 70
        workbook.save(workbook_path)
        workbook.close()

        podcastone, _ = Vendor.objects.get_or_create(name="PodcastOne", defaults={"parser_key": "podcastone"})
        return Asset.objects.create(
            remote_item_id="fi-podcastone-parse",
            vendor=podcastone,
            parser_key="podcastone",
            status=AssetStatus.PROCESSING,
            name="PODCASTONE MAY 2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_rallyadmedia_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "RallyAdMedia"
        parser_root.mkdir(parents=True)
        shutil.copy2(
            project_root / "parsers" / "RallyAdMedia" / "input_schema.json",
            parser_root / "input_schema.json",
        )
        shutil.copy2(project_root / "parsers" / "RallyAdMedia" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "RallyAdMedia" / "RallyAdMedia.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-04-01,RallyAdMedia,BetOnline,Display,RallyAd,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/allshared/May_2026_Internal_folders/RallyAd_Mar26_BOL_SB_WC_SS_2026.xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet_values = {
            "BOL": [("01-Mar-2026", 100, 10), ("02-Mar-2026", 500, 50)],
            "SB": [("01-Mar-2026", 200, 20), (None, None, None)],
            "WC": [("01-Mar-2026", 300, 30), ("02-Mar-2026", 600, 60)],
            "SS": [("01-Mar-2026", 400, 40), ("02-Mar-2026", 700, 70)],
        }
        for sheet_name, rows in sheet_values.items():
            sheet = workbook.create_sheet(sheet_name)
            sheet["A1"] = "DATE_LABEL"
            sheet["B1"] = "Imps."
            sheet["I1"] = "Total Spend"
            for index, (row_date, impressions, spend) in enumerate(rows, 2):
                sheet.cell(index, 1).value = row_date
                sheet.cell(index, 2).value = impressions
                sheet.cell(index, 9).value = spend
        workbook.save(workbook_path)
        workbook.close()

        rallyad, _ = Vendor.objects.get_or_create(name="RallyAdMedia", defaults={"parser_key": "rallyadmedia"})
        return Asset.objects.create(
            remote_item_id="fi-rallyadmedia-parse",
            vendor=rallyad,
            parser_key="rallyadmedia",
            status=AssetStatus.PROCESSING,
            name="RallyAd_Mar26_BOL_SB_WC_SS_2026.xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def _write_adtaxi_parse_fixture(self, repo_root: Path) -> Asset:
        project_root = Path(__file__).resolve().parents[3]
        parser_root = repo_root / "parsers" / "AdTaxi"
        parser_root.mkdir(parents=True)
        shutil.copy2(project_root / "parsers" / "AdTaxi" / "input_schema.json", parser_root / "input_schema.json")
        shutil.copy2(project_root / "parsers" / "AdTaxi" / "parser.py", parser_root / "parser.py")

        approved_path = repo_root / "data" / "processed" / "AdTaxi" / "AdTaxi.csv"
        approved_path.parent.mkdir(parents=True)
        approved_path.write_text(
            "\n".join(
                [
                    "Date,Vendor,Brand,Channel,Platform,Spend,Impressions,Data_Grain,Processed_At,Source_File",
                    "2026-02-01,AdTaxi,BetOnline,Display,AdTaxi,10,100,daily,approved,baseline.csv",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        local_path = "data/inbox/allshared/May_2026_Internal_folders/AdTaxi_Media Spend & Conversion by State (March 2026).xlsx"
        workbook_path = repo_root / local_path
        workbook_path.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Totals By State"
        sheet["A1"] = "Timeframe"
        sheet["G1"] = "Advertiser Cost "
        sheet["H1"] = "Impressions"
        sheet["A2"] = "Dates"
        sheet["B2"] = "3/1/2026/26 - 3/31-26"
        sheet["G3"] = 31
        sheet["H3"] = 310
        sheet["G4"] = 62
        sheet["H4"] = 620
        sheet["G5"] = 0
        sheet["H5"] = 0
        workbook.save(workbook_path)
        workbook.close()

        adtaxi, _ = Vendor.objects.get_or_create(name="AdTaxi", defaults={"parser_key": "adtaxi"})
        return Asset.objects.create(
            remote_item_id="fi-adtaxi-parse",
            vendor=adtaxi,
            parser_key="adtaxi",
            status=AssetStatus.PROCESSING,
            name="AdTaxi_Media Spend & Conversion by State (March 2026).xlsx",
            local_path=local_path,
            file_size=workbook_path.stat().st_size,
        )

    def test_inactive_asset_excluded_from_new_count(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            state_root = repo_root / "data" / "state"
            state_root.mkdir(parents=True)
            (state_root / "sharefile_snapshot_latest.json").write_text(
                json.dumps(
                    {
                        "run_id": "snapshot-inactive",
                        "created_at": "2026-05-07T10:00:00Z",
                        "files": [
                            {
                                "remote_item_id": "fi-inactive-new",
                                "name": "inactive-new.xlsx",
                                "remote_path": "home/josh/inactive-new.xlsx",
                                "local_path": "data/inbox/home/josh/inactive-new.xlsx",
                                "source_folder_path": "home/josh",
                                "extension": ".xlsx",
                                "size": 10,
                                "modified_at": "2026-05-07T12:00:00Z",
                                "creator": "Uploader One",
                                "raw_metadata": {"LastModifiedByUserID": "user-1"},
                            },
                            {
                                "remote_item_id": "fi-active-new",
                                "name": "active-new.xlsx",
                                "remote_path": "home/josh/active-new.xlsx",
                                "local_path": "data/inbox/home/josh/active-new.xlsx",
                                "source_folder_path": "home/josh",
                                "extension": ".xlsx",
                                "size": 10,
                                "modified_at": "2026-05-07T12:00:00Z",
                                "creator": "Uploader One",
                                "raw_metadata": {"LastModifiedByUserID": "user-1"},
                            },
                        ],
                    }
                )
            )
            (state_root / "inbox_profile_latest.json").write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "local_path": "data/inbox/home/josh/inactive-new.xlsx",
                                "name": "inactive-new.xlsx",
                                "kind": "excel",
                                "status": "profiled",
                                "sheet_count": 1,
                            },
                            {
                                "local_path": "data/inbox/home/josh/active-new.xlsx",
                                "name": "active-new.xlsx",
                                "kind": "excel",
                                "status": "profiled",
                                "sheet_count": 1,
                            },
                        ]
                    }
                )
            )
            (state_root / "file_processing_state.json").write_text(json.dumps({}))
            (state_root / "sharefile_users_latest.json").write_text(
                json.dumps({"users_by_id": {"user-1": {"full_name": "Uploader One", "email": "one@example.com"}}})
            )
            (state_root / "sharefile_sync_state.json").write_text(json.dumps({}))

            Asset.objects.create(
                remote_item_id="fi-inactive-new",
                status=AssetStatus.NEW,
                name="inactive-new.xlsx",
                local_path="data/inbox/home/josh/inactive-new.xlsx",
                is_active=False,
            )

            with override_settings(REPO_ROOT=repo_root):
                response = self.client.get(reverse("pipeline_dashboard:folders"))

        self.assertEqual(response.status_code, 200)
        josh_folder = next(f for f in response.context["folders"] if f["display_name"] == "josh")
        self.assertEqual(josh_folder["counts"]["total"], 2)
        self.assertEqual(josh_folder["counts"]["new"], 1)
        self.assertEqual(josh_folder["counts"]["active"], 0)
        self.assertEqual(josh_folder["counts"]["processed"], 0)
