from __future__ import annotations

import csv
import importlib.util
import json
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings

from testifize_pipeline.config import load_dotenv
from testifize_pipeline.sharefile.client import ShareFileClient, ShareFileConfig

from .file_review import ReviewPreviewError, build_file_preview, inbox_file_path
from .models import Asset, ParsedOutput, Vendor


class ParserWorkflowError(ValueError):
    pass


@dataclass(frozen=True)
class ParserPaths:
    vendor_key: str
    parser_root: Path
    schema_path: Path
    parser_path: Path


@dataclass(frozen=True)
class ParsedRows:
    rows: list[dict[str, Any]]
    columns: list[str]
    summary: dict[str, Any]
    approved_path: Path | None


def build_parse_preview(asset: Asset) -> dict[str, Any]:
    if not asset.local_path:
        raise ParserWorkflowError("Asset has no local file path.")
    file_preview = build_file_preview(asset.local_path, asset_preview_metadata(asset))
    validation = validate_asset_parser(asset)
    return {"file": file_preview, "validation": validation}


def build_parse_result_preview(asset: Asset, sheet_name: str | None = None) -> dict[str, Any]:
    parsed = parse_asset_rows(asset, sheet_name=sheet_name)
    comparison = compare_to_approved(parsed.rows, parsed.approved_path)
    output_label = approval_filename_label(parsed.summary)
    return {
        "candidate": {
            "summary": {**serializable_summary(parsed.summary), "vendor_name": asset.vendor.name if asset.vendor else ""},
            "output_filename": next_output_path(asset.vendor, output_label)[0].name if asset.vendor else "",
        },
        "comparison": comparison,
        "charts": build_chart_payload(parsed.rows, asset.vendor, parsed.summary),
        "parsed_table": parsed_table_payload(parsed),
    }


def _review_file_preview(asset: Asset | None) -> dict[str, Any] | None:
    if not asset:
        return None
    return {
        "name": asset.name,
        "created_at": asset.remote_created_at.isoformat() if asset.remote_created_at else "",
        "uploaded_by": asset.created_by_name or "",
    }


def _review_validation_preview(asset: Asset | None) -> dict[str, Any] | None:
    if not asset:
        return None
    paths = parser_paths_for_vendor(asset.vendor) if asset.vendor else None
    return {
        "ok": True,
        "vendor": asset.vendor.name if asset.vendor else "",
        "parser_key": paths.vendor_key if paths else "",
        "parser_path": display_path(paths.parser_path) if paths else "",
        "schema_path": display_path(paths.schema_path) if paths else "",
    }


def build_review_payload(parsed: ParsedOutput) -> dict[str, Any]:
    """Reconstruct the parse-review payload from a saved ParsedOutput CSV."""
    candidates = []
    if parsed.output_path:
        candidates.append(settings.REPO_ROOT / parsed.output_path)
    if parsed.asset and parsed.asset.output_path:
        candidates.append(settings.REPO_ROOT / parsed.asset.output_path)

    source_path = None
    for candidate in candidates:
        if candidate.exists():
            source_path = candidate
            break

    if source_path is None:
        raise ParserWorkflowError("Parsed CSV file not found.")

    with source_path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        columns = list(reader.fieldnames or [])
        rows = list(reader)

    if not columns:
        raise ParserWorkflowError("Parsed CSV has no columns.")

    summary = summarize_rows(rows)
    approved_path = None
    if parsed.approved_path:
        approved_path = settings.REPO_ROOT / parsed.approved_path
        if not approved_path.exists():
            approved_path = None

    preview_rows = rows[:200]
    return {
        "candidate": {
            "summary": {**serializable_summary(summary), "vendor_name": parsed.vendor.name if parsed.vendor else ""},
            "output_filename": source_path.name,
        },
        "comparison": compare_to_approved(rows, approved_path),
        "charts": build_chart_payload(rows, parsed.vendor, summary),
        "parsed_table": {
            "columns": columns,
            "rows": [
                [csv_preview_value(row.get(column, "")) for column in columns]
                for row in preview_rows
            ],
            "row_count": len(rows),
            "preview_count": len(preview_rows),
            "truncated": len(rows) > 200,
        },
        "file": _review_file_preview(parsed.asset),
        "validation": _review_validation_preview(parsed.asset),
    }


def parsed_table_payload(parsed: ParsedRows, limit: int = 200) -> dict[str, Any]:
    preview_rows = parsed.rows[:limit]
    return {
        "columns": parsed.columns,
        "rows": [
            [csv_preview_value(row.get(column, "")) for column in parsed.columns]
            for row in preview_rows
        ],
        "row_count": len(parsed.rows),
        "preview_count": len(preview_rows),
        "truncated": len(parsed.rows) > limit,
    }


def csv_preview_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def validate_asset_parser(asset: Asset) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    schema: dict[str, Any] = {}
    paths = parser_paths_for_vendor(asset.vendor)

    if not asset.vendor:
        errors.append("No vendor is assigned.")
    if not paths:
        errors.append("No parser folder is configured for this vendor.")
        return validation_payload(asset, None, schema, errors, warnings)

    if not paths.schema_path.exists():
        errors.append(f"Missing input schema: {display_path(paths.schema_path)}")
    else:
        try:
            schema = json.loads(paths.schema_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            errors.append(f"Input schema is not valid JSON: {exc}")

    if not paths.parser_path.exists():
        errors.append(f"Missing parser file: {display_path(paths.parser_path)}")

    if not asset.local_path:
        errors.append("Asset has no local file path.")
    else:
        try:
            source_path = inbox_file_path(asset.local_path)
        except ReviewPreviewError as exc:
            errors.append(str(exc))
        else:
            if schema:
                errors.extend(validate_schema_against_file(source_path, schema))

    approved_path = approved_csv_path(asset.vendor)
    if not approved_path:
        warnings.append("No approved historical CSV exists for comparison.")

    return validation_payload(asset, paths, schema, errors, warnings, approved_path)


def probe_sheet_validation(asset: Asset, sheet_name: str) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    schema: dict[str, Any] = {}
    paths = parser_paths_for_vendor(asset.vendor)

    if not asset.vendor:
        errors.append("No vendor is assigned.")
    if not paths:
        errors.append("No parser folder is configured for this vendor.")
        return validation_payload(asset, None, schema, errors, warnings)

    if not paths.schema_path.exists():
        errors.append(f"Missing input schema: {display_path(paths.schema_path)}")
    else:
        try:
            schema = json.loads(paths.schema_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            errors.append(f"Input schema is not valid JSON: {exc}")

    if not paths.parser_path.exists():
        errors.append(f"Missing parser file: {display_path(paths.parser_path)}")

    if not asset.local_path:
        errors.append("Asset has no local file path.")
    else:
        try:
            source_path = inbox_file_path(asset.local_path)
        except ReviewPreviewError as exc:
            errors.append(str(exc))
        else:
            if schema:
                file_type = schema.get("file_type", "").lower()
                extension = source_path.suffix.lower().lstrip(".")
                if file_type and file_type != extension:
                    if not (file_type == "xlsx" and extension == "xlsm"):
                        errors.append(f"Schema expects {file_type}; selected file is {extension or 'unknown'}.")
                if file_type in {"xlsx", "xlsm"} or extension in {"xlsx", "xlsm"}:
                    errors.extend(validate_excel_schema_probe(source_path, schema, sheet_name))
                else:
                    errors.append("Sheet probing is only supported for Excel files.")

    approved_path = approved_csv_path(asset.vendor)
    if not approved_path:
        warnings.append("No approved historical CSV exists for comparison.")

    return validation_payload(asset, paths, schema, errors, warnings, approved_path)


def parse_asset_rows(asset: Asset, sheet_name: str | None = None) -> ParsedRows:
    if not asset.vendor:
        raise ParserWorkflowError("No vendor is assigned.")
    paths = parser_paths_for_vendor(asset.vendor)
    if not paths:
        raise ParserWorkflowError("No parser folder is configured for this vendor.")
    if not paths.schema_path.exists():
        raise ParserWorkflowError(f"Missing input schema: {display_path(paths.schema_path)}")
    if not paths.parser_path.exists():
        raise ParserWorkflowError(f"Missing parser file: {display_path(paths.parser_path)}")

    source_path = inbox_file_path(asset.local_path)
    schema = load_json(paths.schema_path)

    if sheet_name:
        errors: list[str] = []
        file_type = schema.get("file_type", "").lower()
        extension = source_path.suffix.lower().lstrip(".")
        if file_type and file_type != extension:
            if not (file_type == "xlsx" and extension == "xlsm"):
                errors.append(f"Schema expects {file_type}; selected file is {extension or 'unknown'}.")
        if file_type in {"xlsx", "xlsm"} or extension in {"xlsx", "xlsm"}:
            errors.extend(validate_excel_schema_probe(source_path, schema, sheet_name))
        else:
            errors.append("Sheet probing is only supported for Excel files.")
        if errors:
            raise ParserWorkflowError("; ".join(errors))
    else:
        validation = validate_asset_parser(asset)
        if not validation["ok"]:
            raise ParserWorkflowError("; ".join(validation["errors"]))

    approved_path = approved_csv_path(asset.vendor)
    output_columns = load_output_columns(approved_path, schema)
    module = load_parser_module(paths.parser_path)
    if not hasattr(module, "parse_file"):
        raise ParserWorkflowError(f"{display_path(paths.parser_path)} does not expose parse_file().")

    rows = module.parse_file(source_path, schema, output_columns, sheet_name=sheet_name)
    if not rows:
        raise ParserWorkflowError("Parser returned no rows.")
    normalized_rows = [{column: row.get(column, "") for column in output_columns} for row in rows]
    summary = summarize_rows(normalized_rows)
    return ParsedRows(rows=normalized_rows, columns=output_columns, summary=summary, approved_path=approved_path)


def approve_asset_parser(asset: Asset, client: ShareFileClient | None = None) -> ParsedOutput:
    parsed = stage_asset_parser(asset)
    upload_item = upload_approved_output(settings.REPO_ROOT / parsed.output_path, parsed.vendor, parsed.comparison_summary, client=client)
    parsed.comparison_summary = {
        **(parsed.comparison_summary or {}),
        "sharefile_item_id": upload_item.id,
        "sharefile_filename": upload_item.name,
    }
    parsed.comparison_status = "sent_for_approval"
    parsed.save(update_fields=["comparison_status", "comparison_summary"])
    return parsed


def stage_asset_parser(asset: Asset) -> ParsedOutput:
    parsed_rows = parse_asset_rows(asset)
    if not asset.vendor:
        raise ParserWorkflowError("No vendor is assigned.")
    summary = parsed_rows.summary
    output_path, version = next_output_path(asset.vendor, approval_filename_label(summary))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_rows(output_path, parsed_rows.columns, parsed_rows.rows)

    comparison = {
        **compare_to_approved(parsed_rows.rows, parsed_rows.approved_path),
        "approval_folder_label": approval_folder_label(summary),
        "approval_filename_label": approval_filename_label(summary),
    }
    parsed = ParsedOutput.objects.create(
        asset=asset,
        vendor=asset.vendor,
        output_path=relative_path(output_path),
        approved_path=relative_path(parsed_rows.approved_path) if parsed_rows.approved_path else "",
        reporting_period=summary["period_label"],
        period_start=summary["period_start"],
        period_end=summary["period_end"],
        version=version,
        row_count=summary["row_count"],
        total_spend=summary["total_spend"],
        total_impressions=summary["total_impressions"],
        comparison_status=comparison["status"],
        comparison_summary=comparison,
    )
    return parsed


def upload_approved_output(
    output_path: Path,
    vendor: Vendor,
    summary: dict[str, Any],
    client: ShareFileClient | None = None,
):
    root_id = approval_root_id()
    approval_month = approval_folder_label(summary)
    client = client or build_sharefile_client()
    folder = client.ensure_folder_path(
        root_id, ["Approval", approval_month, vendor.name], copy_access_controls=True
    )
    return client.upload_bytes(
        folder.id,
        output_path.name,
        output_path.read_bytes(),
        content_type="text/csv",
        notify=True,
        overwrite=False,
    )


def finalize_approved_output(parsed_output: ParsedOutput, client: ShareFileClient | None = None):
    final_path = promote_final_output_file(parsed_output)
    root_id = final_root_id()
    period_label = final_period_label(parsed_output)
    client = client or build_sharefile_client()
    try:
        folder = client.ensure_folder_path(root_id, ["Final", period_label], copy_access_controls=True)
        return client.upload_bytes(
            folder.id,
            final_path.name,
            final_path.read_bytes(),
            content_type="text/csv",
            notify=True,
            overwrite=False,
        )
    except Exception:
        final_path.unlink(missing_ok=True)
        raise


def approval_root_id() -> str:
    env = load_dotenv(settings.REPO_ROOT / ".env")
    root_id = env.get("SHAREFILE_APPROVAL_ROOT_ID") or env.get("SHAREFILE_APPROVAL_FOLDER_ID")
    return root_id or "allshared"


def final_root_id() -> str:
    env = load_dotenv(settings.REPO_ROOT / ".env")
    root_id = (
        env.get("SHAREFILE_FINAL_ROOT_ID")
        or env.get("SHAREFILE_FINAL_FOLDER_ID")
        or env.get("SHAREFILE_APPROVAL_ROOT_ID")
        or env.get("SHAREFILE_APPROVAL_FOLDER_ID")
    )
    return root_id or "allshared"


def build_sharefile_client() -> ShareFileClient:
    env = load_dotenv(settings.REPO_ROOT / ".env")
    client = ShareFileClient(ShareFileConfig.from_env(env))
    client.authenticate()
    return client


def parser_paths_for_vendor(vendor: Vendor | None) -> ParserPaths | None:
    if not vendor:
        return None
    candidates = [vendor.name]
    if vendor.parser_key and vendor.parser_key not in candidates:
        candidates.append(vendor.parser_key)
    for key in candidates:
        parser_root = settings.REPO_ROOT / "parsers" / key
        if parser_root.exists():
            return ParserPaths(
                vendor_key=key,
                parser_root=parser_root,
                schema_path=parser_root / "input_schema.json",
                parser_path=parser_root / "parser.py",
            )
    key = vendor.parser_key or vendor.name
    parser_root = settings.REPO_ROOT / "parsers" / key
    return ParserPaths(
        vendor_key=key,
        parser_root=parser_root,
        schema_path=parser_root / "input_schema.json",
        parser_path=parser_root / "parser.py",
    )


def validate_schema_against_file(source_path: Path, schema: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    file_type = schema.get("file_type", "").lower()
    extension = source_path.suffix.lower().lstrip(".")
    if file_type and file_type != extension:
        if not (file_type == "xlsx" and extension == "xlsm"):
            errors.append(f"Schema expects {file_type}; selected file is {extension or 'unknown'}.")

    if file_type in {"xlsx", "xlsm"} or extension in {"xlsx", "xlsm"}:
        errors.extend(validate_excel_schema(source_path, schema))
    elif file_type == "csv" or extension == "csv":
        errors.extend(validate_csv_schema(source_path, schema))
    else:
        errors.append("This file type is not supported by parser validation yet.")
    return errors


def validate_excel_schema_probe(source_path: Path, schema: dict[str, Any], probe_sheet_name: str) -> list[str]:
    from openpyxl import load_workbook

    errors: list[str] = []
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if probe_sheet_name not in workbook.sheetnames:
            return [f"Sheet {probe_sheet_name!r} was not found."]

        worksheets = schema.get("worksheets") or []
        if worksheets:
            worksheet = next(
                (entry for entry in worksheets if entry.get("name") == probe_sheet_name),
                None,
            )
            if worksheet is None:
                declared = [entry.get("name") for entry in worksheets if entry.get("name")]
                if declared:
                    return [
                        f"Sheet {probe_sheet_name!r} is not declared in the input schema; "
                        f"expected one of: {', '.join(declared)}."
                    ]
                return [f"Sheet {probe_sheet_name!r} is not declared in the input schema."]

            header_row = worksheet.get("header_row")
            if not header_row:
                return [f"Input schema worksheet {probe_sheet_name!r} is missing header_row."]

            selected_columns = schema.get("selected_columns") or {}
            sheet = workbook[probe_sheet_name]
            for field, letter in (worksheet.get("columns") or {}).items():
                expected = selected_columns.get(field)
                if not expected:
                    continue
                column_index = column_letter_to_index(letter)
                actual = sheet.cell(header_row, column_index).value
                if actual != expected:
                    errors.append(
                        f"Header mismatch in {probe_sheet_name!r} at {letter}{header_row}: "
                        f"expected {expected!r}, found {actual!r}."
                    )
            return errors

        header = schema.get("header") or {}
        header_row = header.get("row")
        if not header_row:
            return ["Input schema is missing header.row."]

        sheet = workbook[probe_sheet_name]
        for column in header.get("columns", []):
            column_index = column_letter_to_index(column.get("letter", ""))
            expected = column.get("name")
            actual = sheet.cell(header_row, column_index).value
            if actual != expected:
                errors.append(
                    f"Header mismatch at {column.get('letter')}{header_row}: expected {expected!r}, found {actual!r}."
                )
    finally:
        workbook.close()
    return errors


def validate_excel_schema(source_path: Path, schema: dict[str, Any]) -> list[str]:
    from openpyxl import load_workbook

    errors: list[str] = []
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheets = schema.get("worksheets") or []
        if worksheets:
            selected_columns = schema.get("selected_columns") or {}
            for worksheet in worksheets:
                sheet_name = worksheet.get("name")
                if not sheet_name:
                    errors.append("Input schema worksheet is missing name.")
                    continue
                if sheet_name not in workbook.sheetnames:
                    errors.append(f"Sheet {sheet_name!r} was not found.")
                    continue
                header_row = worksheet.get("header_row")
                if not header_row:
                    errors.append(f"Input schema worksheet {sheet_name!r} is missing header_row.")
                    continue
                sheet = workbook[sheet_name]
                for field, letter in (worksheet.get("columns") or {}).items():
                    expected = selected_columns.get(field)
                    if not expected:
                        continue
                    column_index = column_letter_to_index(letter)
                    actual = sheet.cell(header_row, column_index).value
                    if actual != expected:
                        errors.append(
                            f"Header mismatch in {sheet_name!r} at {letter}{header_row}: "
                            f"expected {expected!r}, found {actual!r}."
                        )
            return errors

        sheet_name = schema.get("sheet_name")
        if not sheet_name:
            return ["Input schema is missing sheet_name."]
        header = schema.get("header") or {}
        header_row = header.get("row")
        if not header_row:
            return ["Input schema is missing header.row."]

        if sheet_name not in workbook.sheetnames:
            return [f"Sheet {sheet_name!r} was not found."]
        sheet = workbook[sheet_name]
        for column in header.get("columns", []):
            column_index = column_letter_to_index(column.get("letter", ""))
            expected = column.get("name")
            actual = sheet.cell(header_row, column_index).value
            if actual != expected:
                errors.append(
                    f"Header mismatch at {column.get('letter')}{header_row}: expected {expected!r}, found {actual!r}."
                )
    finally:
        workbook.close()
    return errors


def validate_csv_schema(source_path: Path, schema: dict[str, Any]) -> list[str]:
    required = schema.get("required_columns") or list(schema.get("selected_columns", {}).values())
    encoding = schema.get("encoding") or "utf-8-sig"
    with source_path.open(newline="", encoding=encoding) as file:
        reader = csv.reader(file)
        header = next(reader, [])
    missing = sorted(set(required) - set(header))
    if missing:
        return [f"Missing required CSV columns: {', '.join(missing)}"]
    return []


def approved_csv_path(vendor: Vendor | None) -> Path | None:
    paths = approved_csv_paths(vendor)
    return paths[0] if paths else None


def approved_csv_paths(vendor: Vendor | None) -> list[Path]:
    if not vendor:
        return []
    paths: list[Path] = []
    processed_root = settings.REPO_ROOT / "data" / "processed" / vendor.name
    if processed_root.exists():
        paths.extend(sorted(path for path in processed_root.glob("*.csv") if path.is_file()))
    old_final = settings.REPO_ROOT / "_old" / "final" / f"{vendor.name}.csv"
    if old_final.exists():
        paths.append(old_final)
    return paths


def monthly_processed_output_path(vendor: Vendor, at: datetime | None = None) -> Path:
    month_label = (at or datetime.now()).strftime("%B_%Y")
    return settings.REPO_ROOT / "data" / "processed" / vendor.name / f"{vendor.name}_{month_label}.csv"


def promote_parsed_output_file(parsed_output: ParsedOutput) -> Path:
    if not parsed_output.vendor:
        raise ParserWorkflowError("Parsed output has no vendor.")
    source_path = settings.REPO_ROOT / parsed_output.output_path
    if not source_path.exists():
        raise ParserWorkflowError(f"Parsed CSV was not found: {display_path(source_path)}")
    destination_path = monthly_processed_output_path(parsed_output.vendor)
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    return destination_path


def promote_final_output_file(parsed_output: ParsedOutput) -> Path:
    if not parsed_output.vendor:
        raise ParserWorkflowError("Parsed output has no vendor.")
    source_path = settings.REPO_ROOT / parsed_output.output_path
    if not source_path.exists():
        raise ParserWorkflowError(f"Parsed CSV was not found: {display_path(source_path)}")
    destination_path = final_processed_output_path(parsed_output)
    if destination_path.exists():
        raise ParserWorkflowError(f"Final CSV already exists: {display_path(destination_path)}")
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    return destination_path


def final_processed_output_path(parsed_output: ParsedOutput) -> Path:
    if not parsed_output.vendor:
        raise ParserWorkflowError("Parsed output has no vendor.")
    period_label = final_period_label(parsed_output)
    filename = f"{parsed_output.vendor.name}_{period_label}.csv"
    return settings.REPO_ROOT / "data" / "processed" / parsed_output.vendor.name / filename


def final_period_label(parsed_output: ParsedOutput) -> str:
    if parsed_output.reporting_period:
        return safe_period_label(parsed_output.reporting_period)
    if parsed_output.period_start:
        return parsed_output.period_start.strftime("%B_%Y")
    return approval_period_label()


def approval_folder_label(summary: dict[str, Any] | None = None, at: datetime | None = None) -> str:
    if summary and summary.get("approval_folder_label"):
        return str(summary["approval_folder_label"])
    return approval_period_date(summary, at).strftime("%B_%Y")


def approval_filename_label(summary: dict[str, Any] | None = None, at: datetime | None = None) -> str:
    if summary and summary.get("approval_filename_label"):
        return str(summary["approval_filename_label"])
    return approval_period_date(summary, at).strftime("%b_%Y")


def approval_period_date(summary: dict[str, Any] | None = None, at: datetime | None = None) -> date:
    period_end = summary_period_date(summary, "period_end")
    period_start = summary_period_date(summary, "period_start")
    anchor = period_end or period_start
    if anchor:
        return next_month_start(anchor)
    fallback = at or datetime.now()
    return date(fallback.year, fallback.month, 1)


def summary_period_date(summary: dict[str, Any] | None, key: str) -> date | None:
    if not summary:
        return None
    value = summary.get(key)
    if not value and isinstance(summary.get("generated"), dict):
        value = summary["generated"].get(key)
    return coerce_date(value)


def coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return parse_iso_date(value)
    return None


def next_month_start(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def safe_period_label(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_") or "Unknown_Period"


def load_output_columns(approved_path: Path | None, schema: dict[str, Any]) -> list[str]:
    if approved_path and approved_path.exists():
        with approved_path.open(newline="", encoding="utf-8") as file:
            return next(csv.reader(file))
    defaults = schema.get("output_defaults") or {}
    selected = schema.get("selected_columns") or {}
    columns = ["Date", "Vendor", "Brand"]
    for optional in ["Campaign", "Channel", "Platform", "Marketing_Channel", "Sub_Channel"]:
        if optional in defaults or optional.lower() in selected:
            columns.append(optional)
    columns.extend(["Spend", "Impressions", "Data_Grain", "Processed_At", "Source_File"])
    return list(dict.fromkeys(columns))


def load_parser_module(parser_path: Path):
    spec = importlib.util.spec_from_file_location(f"testifize_vendor_parser_{parser_path.parent.name}", parser_path)
    if not spec or not spec.loader:
        raise ParserWorkflowError(f"Could not load parser: {display_path(parser_path)}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def next_output_path(vendor: Vendor, period_label: str) -> tuple[Path, int]:
    output_root = settings.REPO_ROOT / "data" / "output" / vendor.name
    safe_period = re.sub(r"[^A-Za-z0-9]+", "_", period_label).strip("_") or "Unknown_Period"
    version = 1
    while True:
        candidate = output_root / f"{vendor.name}_{safe_period}_v{version}.csv"
        if not candidate.exists():
            return candidate, version
        version += 1


def approval_period_label(at: datetime | None = None) -> str:
    return approval_folder_label(at=at)


def write_rows(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    dates = []
    for row in rows:
        row_date = parse_iso_date(row.get("Date"))
        if row_date:
            dates.append(row_date)
    period_start = min(dates) if dates else None
    period_end = max(dates) if dates else None
    return {
        "row_count": len(rows),
        "period_start": period_start,
        "period_end": period_end,
        "period_label": period_label(period_start, period_end),
        "total_spend": sum((decimal_value(row.get(spend_column(row), 0)) for row in rows), Decimal("0")),
        "total_impressions": sum((decimal_value(row.get(impression_column(row), 0)) for row in rows), Decimal("0")),
    }


def compare_to_approved(rows: list[dict[str, Any]], approved_path: Path | None) -> dict[str, Any]:
    generated = summarize_rows(rows)
    if not approved_path or not approved_path.exists():
        return {"status": "no_approved_history", "generated": serializable_summary(generated)}

    approved_rows = load_approved_rows(approved_path, generated["period_start"], generated["period_end"])
    approved = summarize_rows(approved_rows)
    spend_diff = generated["total_spend"] - approved["total_spend"]
    impression_diff = generated["total_impressions"] - approved["total_impressions"]
    generated_dates = {row.get("Date") for row in rows if row.get("Date")}
    approved_dates = {row.get("Date") for row in approved_rows if row.get("Date")}
    status = "ok"
    if not approved_rows:
        status = "no_matching_history"
    elif generated_dates != approved_dates or spend_diff != 0 or impression_diff != 0:
        status = "diff"
    return {
        "status": status,
        "approved_path": relative_path(approved_path),
        "generated": serializable_summary(generated),
        "approved": serializable_summary(approved),
        "spend_diff": str(spend_diff),
        "impression_diff": str(impression_diff),
        "missing_dates": sorted(approved_dates - generated_dates),
        "extra_dates": sorted(generated_dates - approved_dates),
    }


def build_chart_payload(rows: list[dict[str, Any]], vendor: Vendor | None, summary: dict[str, Any]) -> dict[str, Any]:
    candidate = period_series_from_rows(rows, f"Parsed {summary['period_label']}")
    baselines = latest_approved_period_series(vendor, summary["period_start"])
    series = [candidate, *baselines]
    return {
        "modes": ["spend", "impressions", "cpm"],
        "series": series,
        "max_day": max((len(item["points"]) for item in series), default=0),
    }


def latest_approved_period_series(vendor: Vendor | None, before_date: date | None, limit: int = 2, min_days: int = 5) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for path in approved_csv_paths(vendor):
        with path.open(newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                row_date = parse_iso_date(row.get("Date"))
                if not row_date:
                    continue
                if before_date and row_date >= before_date.replace(day=1):
                    continue
                key = f"{row_date:%Y-%m}"
                grouped.setdefault(key, []).append(row)

    valid = {key: rows for key, rows in grouped.items() if len(rows) >= min_days}
    selected = sorted(valid.items(), key=lambda item: item[0], reverse=True)[:limit]
    return [
        period_series_from_rows(period_rows, datetime.strptime(f"{period_key}-01", "%Y-%m-%d").strftime("%B_%Y"))
        for period_key, period_rows in selected
    ]


def period_series_from_rows(rows: list[dict[str, Any]], label: str) -> dict[str, Any]:
    grouped: dict[date, dict[str, Decimal]] = {}
    for row in rows:
        row_date = parse_iso_date(row.get("Date"))
        if not row_date:
            continue
        totals = grouped.setdefault(row_date, {"spend": Decimal("0"), "impressions": Decimal("0")})
        totals["spend"] += decimal_value(row.get(spend_column(row), 0))
        totals["impressions"] += decimal_value(row.get(impression_column(row), 0))

    points = []
    for index, row_date in enumerate(sorted(grouped), 1):
        spend = grouped[row_date]["spend"]
        impressions = grouped[row_date]["impressions"]
        cost_per_impression = (spend / impressions) if impressions else Decimal("0")
        points.append(
            {
                "day": index,
                "date": row_date.isoformat(),
                "spend": float(spend),
                "impressions": float(impressions),
                "cpm": float(cost_per_impression),
            }
        )
    return {"label": label, "points": points}


def load_approved_rows(path: Path, period_start: date | None, period_end: date | None) -> list[dict[str, str]]:
    rows = []
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            row_date = parse_iso_date(row.get("Date", ""))
            if (period_start or period_end) and not row_date:
                continue
            if period_start and row_date and row_date < period_start:
                continue
            if period_end and row_date and row_date > period_end:
                continue
            rows.append(row)
    return rows


def _schema_sheet_names(schema: dict[str, Any]) -> list[str]:
    names: list[str] = []
    if schema.get("sheet_name"):
        names.append(schema["sheet_name"])
    for ws in schema.get("worksheets") or []:
        n = ws.get("name")
        if n:
            names.append(n)
    return names


def validation_payload(
    asset: Asset,
    paths: ParserPaths | None,
    schema: dict[str, Any],
    errors: list[str],
    warnings: list[str],
    approved_path: Path | None = None,
) -> dict[str, Any]:
    return {
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "vendor": asset.vendor.name if asset.vendor else "",
        "vendor_id": asset.vendor_id if asset.vendor else None,
        "parser_key": paths.vendor_key if paths else "",
        "schema_path": display_path(paths.schema_path) if paths else "",
        "parser_path": display_path(paths.parser_path) if paths else "",
        "approved_path": display_path(approved_path) if approved_path else "",
        "sheet_name": schema.get("sheet_name", ""),
        "schema_sheet_names": _schema_sheet_names(schema),
        "header_row": (schema.get("header") or {}).get("row", ""),
        "output_defaults": schema.get("output_defaults") or {},
    }


def asset_preview_metadata(asset: Asset) -> dict[str, Any]:
    return {
        "name": asset.name,
        "source_folder_path": asset.source_folder_label,
        "local_path": asset.local_path,
        "remote_path": asset.remote_path,
        "remote_item_id": asset.remote_item_id,
        "extension": Path(asset.local_path).suffix.lower() if asset.local_path else "",
        "size": asset.file_size or 0,
        "created_at": asset.remote_created_at.isoformat() if asset.remote_created_at else "",
        "modified_at": asset.remote_modified_at.isoformat() if asset.remote_modified_at else "",
        "uploaded_by": asset.created_by_name,
        "uploader_email": asset.created_by_email,
    }


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        if not char.isalpha():
            continue
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def parse_iso_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            return None


def period_label(period_start: date | None, period_end: date | None) -> str:
    if not period_start:
        return "Unknown Period"
    if period_end and (period_start.year != period_end.year or period_start.month != period_end.month):
        return f"{period_start:%b_%Y}_to_{period_end:%b_%Y}"
    return f"{period_start:%B_%Y}"


def spend_column(row: dict[str, Any]) -> str:
    return "Daily_Spend" if "Daily_Spend" in row else "Spend"


def impression_column(row: dict[str, Any]) -> str:
    return "Daily_Impressions" if "Daily_Impressions" in row else "Impressions"


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    text = str(value).strip().replace(",", "").replace("$", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def serializable_summary(summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "row_count": summary["row_count"],
        "period_start": summary["period_start"].isoformat() if summary["period_start"] else "",
        "period_end": summary["period_end"].isoformat() if summary["period_end"] else "",
        "period_label": summary["period_label"],
        "total_spend": str(summary["total_spend"]),
        "total_impressions": str(summary["total_impressions"]),
    }


def relative_path(path: Path | None) -> str:
    if not path:
        return ""
    try:
        return str(path.relative_to(settings.REPO_ROOT))
    except ValueError:
        return str(path)


def display_path(path: Path | None) -> str:
    return relative_path(path)
