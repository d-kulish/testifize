from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from django.conf import settings
from openpyxl import load_workbook


EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
CSV_EXTENSIONS = {".csv"}
LEGACY_EXCEL_EXTENSIONS = {".xls"}
MAX_PREVIEW_ROWS = 40
MAX_PREVIEW_COLUMNS = 30


class ReviewPreviewError(ValueError):
    pass


def build_file_preview(local_path: str, metadata: dict[str, Any]) -> dict[str, Any]:
    path = inbox_file_path(local_path)
    extension = path.suffix.lower()
    details = {
        "name": metadata.get("name") or path.name,
        "folder": metadata.get("source_folder_path") or str(path.parent),
        "local_path": local_path,
        "remote_path": metadata.get("remote_path") or "",
        "remote_item_id": metadata.get("remote_item_id") or "",
        "extension": extension,
        "size": metadata.get("size") or path.stat().st_size,
        "created_at": metadata.get("created_at") or "",
        "modified_at": metadata.get("modified_at") or "",
        "uploaded_by": metadata.get("uploaded_by") or "",
        "uploader_email": metadata.get("uploader_email") or "",
    }

    if extension in EXCEL_EXTENSIONS:
        sheets = preview_workbook(path)
        return {**details, "kind": "excel", "sheet_count": len(sheets), "sheets": sheets}
    if extension in CSV_EXTENSIONS:
        sheet = preview_csv(path)
        return {**details, "kind": "csv", "sheet_count": 1, "sheets": [sheet]}
    if extension in LEGACY_EXCEL_EXTENSIONS:
        return {
            **details,
            "kind": "legacy_excel",
            "sheet_count": metadata.get("sheet_count") or 0,
            "sheets": [],
            "warning": "Legacy .xls preview is not available yet.",
        }
    return {
        **details,
        "kind": "unsupported",
        "sheet_count": metadata.get("sheet_count") or 0,
        "sheets": [],
        "warning": f"{extension or 'This file type'} cannot be previewed yet.",
    }


def inbox_file_path(local_path: str) -> Path:
    repo_root = Path(settings.REPO_ROOT).resolve()
    inbox_root = (repo_root / "data" / "inbox").resolve()
    path = (repo_root / local_path).resolve()
    try:
        path.relative_to(inbox_root)
    except ValueError as exc:
        raise ReviewPreviewError("File is outside data/inbox.") from exc
    if not path.exists() or not path.is_file():
        raise ReviewPreviewError("Local file is missing.")
    return path


def preview_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [
                [_cell_text(value) for value in row]
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row or MAX_PREVIEW_ROWS, MAX_PREVIEW_ROWS),
                    max_col=min(sheet.max_column or MAX_PREVIEW_COLUMNS, MAX_PREVIEW_COLUMNS),
                    values_only=True,
                )
            ]
            header = header_from_rows(rows)
            sheets.append(
                {
                    "name": sheet_name,
                    "row_count": sheet.max_row or len(rows),
                    "column_count": sheet.max_column or max((len(row) for row in rows), default=0),
                    "header_row_index": header["row_index"],
                    "headers": header["values"],
                    "rows": rows,
                }
            )
        return sheets
    finally:
        workbook.close()


def preview_csv(path: Path) -> dict[str, Any]:
    rows, encoding = _read_csv_rows(path)
    header = header_from_rows(rows)
    return {
        "name": "CSV",
        "encoding": encoding,
        "row_count": len(rows),
        "column_count": max((len(row) for row in rows), default=0),
        "header_row_index": header["row_index"],
        "headers": header["values"],
        "rows": rows,
    }


def _read_csv_rows(path: Path) -> tuple[list[list[str]], str]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open(newline="", encoding=encoding) as file:
                reader = csv.reader(file)
                return [row[:MAX_PREVIEW_COLUMNS] for _, row in zip(range(MAX_PREVIEW_ROWS), reader)], encoding
        except UnicodeDecodeError:
            continue
    raise ReviewPreviewError("CSV encoding is not readable.")


def header_from_rows(rows: list[list[str]]) -> dict[str, Any]:
    best = {"row_index": None, "values": []}
    best_count = 0
    for index, row in enumerate(rows, 1):
        values = [value.strip() for value in row if value.strip()]
        if len(values) > best_count:
            best = {"row_index": index, "values": values[:MAX_PREVIEW_COLUMNS]}
            best_count = len(values)
        if len(values) >= 3:
            return best
    return best


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    if len(text) > 160:
        return text[:157] + "..."
    return text
