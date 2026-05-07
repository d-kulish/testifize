from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_file(source_path: Path, input_schema: dict[str, Any], output_columns: list[str]) -> list[dict[str, Any]]:
    defaults = input_schema["output_defaults"]
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[input_schema["sheet_name"]]
        daily_totals: dict[str, dict[str, Decimal]] = {}

        for table in input_schema.get("tables", []):
            validate_table_label(sheet, table)
            columns = table["columns"]
            date_column = column_letter_to_index(columns["date"])
            impressions_column = column_letter_to_index(columns["impressions"])
            spend_column = column_letter_to_index(columns["spend"])
            first_data_row = table.get("first_data_row", table["header_row"] + 1)

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=first_data_row, max_row=sheet.max_row, values_only=True),
                start=first_data_row,
            ):
                date_value = value_at(row, date_column)
                impression_value = value_at(row, impressions_column)
                spend_value = value_at(row, spend_column)

                if date_value in (None, ""):
                    continue

                row_date = parse_date(date_value, row_number, table["name"])
                totals = daily_totals.setdefault(row_date, {"Spend": Decimal("0"), "Impressions": Decimal("0")})
                totals["Spend"] += parse_decimal(spend_value, row_number, f"{table['name']} spend")
                totals["Impressions"] += parse_decimal(
                    impression_value,
                    row_number,
                    f"{table['name']} impressions",
                )

        if not daily_totals:
            raise ValueError(f"No rows were parsed from {source_path}.")

        rows: list[dict[str, Any]] = []
        processed_at = datetime.now().isoformat(timespec="seconds")
        for row_date in sorted(daily_totals):
            totals = daily_totals[row_date]
            parsed = {
                "Date": row_date,
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Channel": defaults["Channel"],
                "Platform": defaults["Platform"],
                "Spend": decimal_text(totals["Spend"]),
                "Impressions": decimal_text(totals["Impressions"]),
                "Data_Grain": defaults["Data_Grain"],
                "Processed_At": processed_at,
                "Source_File": source_path.name,
            }
            rows.append({column: parsed.get(column, "") for column in output_columns})

        return rows
    finally:
        workbook.close()


def validate_table_label(sheet: Any, table: dict[str, Any]) -> None:
    label_cell = table.get("label_cell")
    if not label_cell:
        return
    row_number, column_index = split_cell_reference(label_cell)
    row = next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True), ())
    actual = value_at(row, column_index)
    expected = table["name"]
    if actual != expected:
        raise ValueError(f"Expected table label {expected!r} at {label_cell}, found {actual!r}.")


def split_cell_reference(cell_reference: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Za-z]+)([0-9]+)", cell_reference)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_reference!r}")
    return int(match.group(2)), column_letter_to_index(match.group(1))


def value_at(row: tuple[Any, ...], one_based_column: int) -> Any:
    index = one_based_column - 1
    if index >= len(row):
        return None
    return row[index]


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int, table_name: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid {table_name} date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"
