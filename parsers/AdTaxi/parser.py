from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DAILY_SECTION_MARKER = "Daily by State"


def parse_file(source_path: Path, input_schema: dict[str, Any], output_columns: list[str], sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Parse AdTaxi multi-sheet Excel files containing daily spend/impressions by state.

    Each sheet has a dual-table layout: a "Totals by State" table on the left and a
    "Daily by State" table on the right.  The parser scans every sheet for the right
    table, skips the state-level aggregate rows, and collects the actual day-by-day
    rows.  Spend and impressions are aggregated across all sheets and all states.
    """
    defaults = input_schema.get("output_defaults", {})
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        daily_spend: defaultdict[date, Decimal] = defaultdict(lambda: Decimal("0"))
        daily_impressions: defaultdict[date, Decimal] = defaultdict(lambda: Decimal("0"))

        sheets_to_process = [sheet_name] if sheet_name else workbook.sheetnames
        for sn in sheets_to_process:
            if sn not in workbook.sheetnames:
                raise ValueError(f"Sheet {sn!r} not found in {source_path.name}.")
            _extract_daily_rows(workbook[sn], daily_spend, daily_impressions)

        if not daily_spend:
            raise ValueError(f"No daily data was found in {source_path.name}.")

        processed_at = datetime.now().isoformat(timespec="seconds")
        rows: list[dict[str, Any]] = []
        for row_date in sorted(daily_spend):
            parsed = {
                "Date": row_date.isoformat(),
                "Vendor": defaults.get("Vendor", "AdTaxi"),
                "Brand": defaults.get("Brand", "BetOnline"),
                "Channel": defaults.get("Channel", "Display, CTV"),
                "Platform": defaults.get("Platform", "AdTaxi"),
                "Spend": decimal_text(daily_spend[row_date]),
                "Impressions": decimal_text(daily_impressions.get(row_date, Decimal("0"))),
                "Data_Grain": defaults.get("Data_Grain", "daily"),
                "Processed_At": processed_at,
                "Source_File": source_path.name,
            }
            rows.append({column: parsed.get(column, "") for column in output_columns})

        return rows
    finally:
        workbook.close()


def _extract_daily_rows(sheet, daily_spend: defaultdict, daily_impressions: defaultdict) -> None:
    """Scan a single sheet for the 'Daily by State' section and aggregate values."""
    in_daily_section = False

    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 10:
            continue

        col_h = row[7]  # Column H

        # Detect the "Daily by State" header row
        if isinstance(col_h, str) and DAILY_SECTION_MARKER in col_h:
            in_daily_section = True
            continue

        if not in_daily_section:
            continue

        # Skip state total rows and sub-headers in the daily section
        if isinstance(col_h, str):
            continue
        if col_h is None:
            continue

        # Parse the date from the Excel cell
        row_date = _parse_cell_date(col_h)
        if row_date is None:
            continue

        # Spend is in column I (index 8), Impressions in column J (index 9)
        spend = _parse_decimal(row[8])
        impressions = _parse_decimal(row[9])

        daily_spend[row_date] += spend
        daily_impressions[row_date] += impressions


def _parse_cell_date(value: Any) -> date | None:
    """Extract a date from an Excel cell value."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        # Handle verbose formats like "Thu Mar 26 2026 02:00:00 GMT+0200 ..."
        match = re.search(r"\b([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})\b", text)
        if match:
            try:
                return datetime.strptime(
                    f"{match.group(1)} {match.group(2)} {match.group(3)}",
                    "%b %d %Y",
                ).date()
            except ValueError:
                pass
        # Fallback to common slash formats
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(value: Any) -> Decimal:
    """Safely parse a numeric cell value into a Decimal."""
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "").strip("()")
        try:
            parsed = Decimal(cleaned)
            if value.strip().startswith("(") and value.strip().endswith(")"):
                parsed = -parsed
            return parsed
        except (InvalidOperation, ValueError):
            return Decimal("0")
    return Decimal("0")


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"
