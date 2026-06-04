from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_file(source_path: Path, input_schema: dict[str, Any], output_columns: list[str], sheet_name: str | None = None) -> list[dict[str, Any]]:
    defaults = input_schema["output_defaults"]
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        daily_totals: dict[str, dict[str, Decimal]] = {}

        for worksheet in input_schema.get("worksheets", []):
            actual_sheet_name = resolve_sheet_name(workbook, worksheet)
            if actual_sheet_name is None:
                raise ValueError(f"No sheet matching {worksheet.get('match_keywords') or worksheet.get('name')!r} was found.")

            sheet = workbook[actual_sheet_name]
            discovered = discover_columns(sheet, worksheet)
            validate_headers(sheet, worksheet, discovered)

            date_column = column_letter_to_index(discovered["date"])
            impressions_column = column_letter_to_index(discovered["impressions"])
            spend_column = column_letter_to_index(discovered["spend"])

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=worksheet["first_data_row"], max_row=sheet.max_row, values_only=True),
                start=worksheet["first_data_row"],
            ):
                date_value = value_at(row, date_column)
                if date_value in (None, ""):
                    continue

                row_date = parse_date(date_value, row_number, actual_sheet_name)
                totals = daily_totals.setdefault(row_date, {"Spend": Decimal("0"), "Impressions": Decimal("0")})
                totals["Spend"] += parse_decimal(value_at(row, spend_column), row_number, f"{actual_sheet_name} spend")
                totals["Impressions"] += parse_decimal(
                    value_at(row, impressions_column),
                    row_number,
                    f"{actual_sheet_name} impressions",
                )

        if not daily_totals:
            raise ValueError(f"No rows were parsed from {source_path}.")

        processed_at = datetime.now().isoformat(timespec="seconds")
        rows: list[dict[str, Any]] = []
        for row_date in sorted(daily_totals):
            totals = daily_totals[row_date]
            parsed = {
                "Date": row_date,
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Channel": defaults["Channel"],
                "Platform": defaults["Platform"],
                "Spend": decimal_text(totals["Spend"]),
                "Impressions": decimal_text(totals["Impressions"]),
                "Data_Grain": defaults["Data_Grain"],
                "Processed_At": processed_at,
                "Source_File": source_path.name,
            }
            rows.append({column: parsed.get(column, "") for column in output_columns})

        return rows
    finally:
        workbook.close()


def resolve_sheet_name(workbook: Any, worksheet: dict[str, Any]) -> str | None:
    """Find the actual sheet name using match_keywords or exact name."""
    keywords = worksheet.get("match_keywords")
    if keywords:
        keywords_lower = [str(k).lower() for k in keywords]
        for name in workbook.sheetnames:
            name_lower = name.lower()
            if all(kw in name_lower for kw in keywords_lower):
                return name
        return None
    exact = worksheet.get("name")
    if exact and exact in workbook.sheetnames:
        return exact
    return None


def discover_columns(sheet: Any, worksheet: dict[str, Any]) -> dict[str, str]:
    """Return a dict mapping field → column letter.

    Prefers columns_by_header (scan header row for text), falls back to
    the legacy columns dict (already letters).
    """
    columns_by_header = worksheet.get("columns_by_header")
    if not columns_by_header:
        return worksheet.get("columns", {})

    header_row = worksheet.get("header_row")
    if not header_row:
        raise ValueError("Worksheet is missing header_row required for columns_by_header.")

    result: dict[str, str] = {}
    header_values = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

    for field, expected_text in columns_by_header.items():
        expected_clean = str(expected_text).strip()
        found = False
        for idx, cell in enumerate(header_values, start=1):
            if cell is not None and str(cell).strip() == expected_clean:
                result[field] = index_to_column_letter(idx)
                found = True
                break
        if not found:
            raise ValueError(
                f"Header '{expected_clean}' for field '{field}' not found in row {header_row} of sheet '{sheet.title}'."
            )
    return result


def validate_headers(sheet: Any, worksheet: dict[str, Any], discovered: dict[str, str]) -> None:
    header_row = worksheet["header_row"]
    expected = {
        "date": "Day",
        "impressions": "Audio Impressions",
        "spend": "$ By Day",
    }
    for field, expected_name in expected.items():
        column_index = column_letter_to_index(discovered[field])
        actual = next(
            sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                min_col=column_index,
                max_col=column_index,
                values_only=True,
            )
        )[0]
        if actual != expected_name:
            raise ValueError(
                f"Header mismatch in {sheet.title} at row {header_row}: "
                f"expected {expected_name!r}, found {actual!r}."
            )


def value_at(row: tuple[Any, ...], one_based_column: int) -> Any:
    index = one_based_column - 1
    if index >= len(row):
        return None
    return row[index]


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def index_to_column_letter(index: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    result = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(remainder + ord("A")))
    return "".join(reversed(result))


def parse_date(value: object, row_number: int, sheet_name: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid {sheet_name} date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"
