from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_file(source_path: Path, input_schema: dict[str, Any], output_columns: list[str]) -> list[dict[str, Any]]:
    defaults = input_schema["output_defaults"]
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[input_schema["sheet_name"]]
        daily_totals: dict[str, dict[str, Decimal]] = {}

        for table in input_schema.get("tables", []):
            validate_table_header(sheet, table)
            columns = table["columns"]
            date_column = column_letter_to_index(columns["date"])
            impressions_column = column_letter_to_index(columns["impressions"])
            spend_column = column_letter_to_index(columns["spend"])

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=table["first_data_row"], max_row=sheet.max_row, values_only=True),
                start=table["first_data_row"],
            ):
                date_value = value_at(row, date_column)
                if date_value == next_table_name(input_schema, table["name"]):
                    break
                if date_value in (None, ""):
                    continue

                row_date = parse_date(date_value, row_number, table["name"])
                totals = daily_totals.setdefault(row_date, {"Spend": Decimal("0"), "Impressions": Decimal("0")})
                totals["Spend"] += parse_decimal(value_at(row, spend_column), row_number, f"{table['name']} spend")
                totals["Impressions"] += parse_decimal(
                    value_at(row, impressions_column),
                    row_number,
                    f"{table['name']} impressions",
                )

        if not daily_totals:
            raise ValueError(f"No rows were parsed from {source_path}.")

        processed_at = datetime.now().isoformat(timespec="seconds")
        rows: list[dict[str, Any]] = []
        for row_date in sorted(daily_totals):
            totals = daily_totals[row_date]
            parsed = {
                "Date": row_date,
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Channel": defaults["Channel"],
                "Platform": defaults["Platform"],
                "Spend": decimal_text(totals["Spend"]),
                "Impressions": decimal_text(totals["Impressions"]),
                "Data_Grain": defaults["Data_Grain"],
                "Processed_At": processed_at,
                "Source_File": source_path.name,
            }
            rows.append({column: parsed.get(column, "") for column in output_columns})

        return rows
    finally:
        workbook.close()


def validate_table_header(sheet: Any, table: dict[str, Any]) -> None:
    header_row = table["header_row"]
    expected = {
        "date": table["name"],
        "impressions": "Impressions",
        "spend": "Spend",
    }
    for field, expected_name in expected.items():
        column_index = column_letter_to_index(table["columns"][field])
        actual = next(
            sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                min_col=column_index,
                max_col=column_index,
                values_only=True,
            )
        )[0]
        if actual != expected_name:
            raise ValueError(
                f"Header mismatch in {table['name']} at row {header_row}: "
                f"expected {expected_name!r}, found {actual!r}."
            )


def next_table_name(input_schema: dict[str, Any], current_name: str) -> str | None:
    tables = input_schema.get("tables", [])
    for index, table in enumerate(tables):
        if table["name"] == current_name and index + 1 < len(tables):
            return tables[index + 1]["name"]
    return None


def value_at(row: tuple[Any, ...], one_based_column: int) -> Any:
    index = one_based_column - 1
    if index >= len(row):
        return None
    return row[index]


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int, table_name: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid {table_name} date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"
