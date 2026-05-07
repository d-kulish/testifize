from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_file(source_path: Path, input_schema: dict[str, Any], output_columns: list[str]) -> list[dict[str, Any]]:
    selected = input_schema["selected_columns"]
    defaults = input_schema["output_defaults"]
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[input_schema["sheet_name"]]
        header_map = build_header_map(input_schema)
        rows: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=input_schema["header"]["row"] + 1,
                max_row=sheet.max_row,
                values_only=True,
            ),
            start=input_schema["header"]["row"] + 1,
        ):
            date_value = row[header_map[selected["date"]] - 1]
            impression_value = row[header_map[selected["impressions"]] - 1]
            spend_value = row[header_map[selected["spend"]] - 1]

            if date_value in (None, ""):
                continue

            parsed = {
                "Date": parse_date(date_value, row_number),
                "Vendor": defaults["Vendor"],
                "Brand": defaults["Brand"],
                "Channel": defaults["Channel"],
                "Platform": defaults["Platform"],
                "Spend": decimal_text(parse_decimal(spend_value, row_number, selected["spend"])),
                "Impressions": decimal_text(parse_decimal(impression_value, row_number, selected["impressions"])),
                "Data_Grain": defaults["Data_Grain"],
                "Processed_At": datetime.now().isoformat(timespec="seconds"),
                "Source_File": source_path.name,
            }
            rows.append({column: parsed.get(column, "") for column in output_columns})

        if not rows:
            raise ValueError(f"No rows were parsed from {source_path}.")
        return rows
    finally:
        workbook.close()


def build_header_map(input_schema: dict[str, Any]) -> dict[str, int]:
    return {
        column["name"]: column_letter_to_index(column["letter"])
        for column in input_schema["header"]["columns"]
    }


def column_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter.upper():
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def parse_date(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid Date at source row {row_number}: {value!r}")


def parse_decimal(value: object, row_number: int, column_name: str) -> Decimal:
    if value in (None, ""):
        raise ValueError(f"Missing {column_name} at row {row_number}.")

    if isinstance(value, str):
        text = value.strip()
        negative = text.startswith("(") and text.endswith(")")
        cleaned = text.strip("()").replace("$", "").replace(",", "")
    else:
        negative = False
        cleaned = str(value)

    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid {column_name} at row {row_number}: {value!r}") from exc

    if negative:
        parsed = -parsed
    if parsed < 0:
        raise ValueError(f"Negative {column_name} at row {row_number}: {value!r}")
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"
